from flask import Flask, render_template, request, redirect, url_for, session, flash, jsonify, make_response, Response
from werkzeug.security import generate_password_hash, check_password_hash
from functools import wraps
import sqlite3
from datetime import datetime, timedelta
import os
import csv
from io import StringIO
import secrets
import requests
import queue
import threading
import json

# Import email notification functions
try:
    from email_notifications import (
        send_appointment_confirmation,
        send_video_call_reminder,
        send_doctor_appointment_notification,
        send_appointment_accepted_email,
        send_appointment_reminder,
        send_doctor_verification_email,
        EMAIL_ENABLED
    )
    print("[Email] Email notification module loaded")
except ImportError as e:
    print(f"[Email] Warning: Could not import email_notifications: {e}")
    EMAIL_ENABLED = False
    # Define dummy functions if import fails
    def send_appointment_confirmation(*args, **kwargs):
        return False
    def send_video_call_reminder(*args, **kwargs):
        return False
    def send_doctor_appointment_notification(*args, **kwargs):
        return False
    def send_appointment_accepted_email(*args, **kwargs):
        return False
    def send_appointment_reminder(*args, **kwargs):
        return False
    def send_doctor_verification_email(*args, **kwargs):
        return False

# Load .env file if present (optional, for local dev)
try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass

# ==================== FLASK-SOCKETIO ====================
try:
    from flask_socketio import SocketIO, emit, join_room, leave_room
    SOCKETIO_AVAILABLE = True
except ImportError:
    SOCKETIO_AVAILABLE = False

# ==================== ENHANCEMENTS MODULE ====================
from enhancements import enhancements_bp, init_enhancement_tables

# ==================== PAYMENT MODULE ====================
from payment import payment_bp, init_payment_tables

# ==================== APPOINTMENT REMINDER SCHEDULER ====================
try:
    from appointment_reminder_scheduler import start_reminder_scheduler
    REMINDER_SCHEDULER_AVAILABLE = True
    print("[Reminder] Reminder scheduler module loaded")
except ImportError as e:
    print(f"[Reminder] Warning: Could not import reminder scheduler: {e}")
    REMINDER_SCHEDULER_AVAILABLE = False

app = Flask(__name__)
app.secret_key = 'medical-system-secret-key-2026'
app.config['PERMANENT_SESSION_LIFETIME'] = 1800

# ==================== SOCKET.IO SETUP ====================
if SOCKETIO_AVAILABLE:
    socketio = SocketIO(app, cors_allowed_origins="*", async_mode='threading')
else:
    socketio = None

# ==================== SSE NOTIFICATION QUEUES ====================
# Each doctor gets a queue. When an appointment is booked,
# we push to their queue. Their browser listens via SSE.
_doctor_queues = {}
_queues_lock   = threading.Lock()

def get_doctor_queue(user_id):
    with _queues_lock:
        if user_id not in _doctor_queues:
            _doctor_queues[user_id] = queue.Queue(maxsize=20)
        return _doctor_queues[user_id]

def push_notification_to_doctor(user_id, data):
    """Push a notification dict to the doctor's SSE queue."""
    q = get_doctor_queue(user_id)
    try:
        q.put_nowait(data)
    except queue.Full:
        pass  # drop oldest if full

# Register enhancements blueprint
app.register_blueprint(enhancements_bp)

# Initialize enhancement tables
init_enhancement_tables()

# Register payment blueprint
app.register_blueprint(payment_bp)

# Initialize payment tables
init_payment_tables()

# ==================== EMAIL — send_doctor_notification() ====================
import smtplib
import threading
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart

def send_doctor_notification(doctor_email, doctor_name, patient_name,
                              date, time, is_emergency=False):
    """
    Send email notification to doctor when a patient books an appointment.
    Uses Gmail SMTP. Credentials from .env:
        GMAIL_USER     = system sender Gmail address
        GMAIL_PASSWORD = Gmail App Password (16 chars, no spaces)
    """
    gmail_user = os.environ.get('GMAIL_USER', '').strip()
    gmail_pass = os.environ.get('GMAIL_PASSWORD', '').strip()

    print(f"[Email] Attempting to send to {doctor_email} | GMAIL_USER set: {bool(gmail_user)}")

    if not gmail_user or not gmail_pass:
        print("[Email] ❌ Not configured. Set GMAIL_USER and GMAIL_PASSWORD in .env")
        return False

    subject = f"{'🚨 EMERGENCY - ' if is_emergency else ''}New Appointment Booked"

    emergency_text = "⚠️  EMERGENCY APPOINTMENT — Please respond immediately!\n\n" if is_emergency else ""
    body = (
        f"Hello Dr. {doctor_name},\n\n"
        + emergency_text
        + "You have a new appointment booked.\n\n"
        + f"Patient Name : {patient_name}\n"
        + f"Date         : {date}\n"
        + f"Time         : {time}\n\n"
        + "Please login to your dashboard for more details.\n\n"
        + "Regards,\nMedical System"
    )

    def _send():
        try:
            msg = MIMEMultipart()
            msg['Subject'] = subject
            msg['From']    = f"Medical System <{gmail_user}>"
            msg['To']      = doctor_email
            msg.attach(MIMEText(body, 'plain'))

            with smtplib.SMTP('smtp.gmail.com', 587, timeout=15) as server:
                server.ehlo()
                server.starttls()
                server.login(gmail_user, gmail_pass)
                server.sendmail(gmail_user, [doctor_email], msg.as_string())

            print(f"[Email] ✅ Sent to Dr. {doctor_name} <{doctor_email}>")
        except smtplib.SMTPAuthenticationError:
            print("[Email] ❌ Auth failed — check GMAIL_USER / GMAIL_PASSWORD in .env")
        except Exception as e:
            print(f"[Email] ❌ {type(e).__name__}: {e}")

    t = threading.Thread(target=_send, daemon=False)  # daemon=False so thread completes even on reload
    t.start()
    return True


# Backward-compat alias used elsewhere in the file
def send_appointment_email(doctor_email, doctor_name, patient_name, appointment_date,
                           appointment_time, consultation_type, symptoms, is_emergency=False):
    return send_doctor_notification(
        doctor_email=doctor_email,
        doctor_name=doctor_name,
        patient_name=patient_name,
        date=appointment_date,
        time=appointment_time,
        is_emergency=is_emergency
    )

# Database initialization
def init_db():
    conn = sqlite3.connect('medical.db')
    c = conn.cursor()
    
    # Users table
    c.execute('''CREATE TABLE IF NOT EXISTS users (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT NOT NULL,
        email TEXT UNIQUE NOT NULL,
        password TEXT NOT NULL,
        role TEXT NOT NULL,
        phone TEXT,
        address TEXT,
        age INTEGER,
        gender TEXT,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    )''')
    
    # Doctors table
    c.execute('''CREATE TABLE IF NOT EXISTS doctors (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id INTEGER,
        specialization TEXT,
        qualification TEXT,
        experience INTEGER,
        consultation_fee REAL,
        available BOOLEAN DEFAULT 1,
        FOREIGN KEY (user_id) REFERENCES users(id)
    )''')
    
    # Patients table
    c.execute('''CREATE TABLE IF NOT EXISTS patients (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id INTEGER,
        blood_group TEXT,
        medical_history TEXT,
        FOREIGN KEY (user_id) REFERENCES users(id)
    )''')
    
    # Consultations table
    c.execute('''CREATE TABLE IF NOT EXISTS consultations (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        patient_id INTEGER,
        doctor_id INTEGER,
        symptoms TEXT,
        diagnosis TEXT,
        prescription TEXT,
        status TEXT DEFAULT 'pending',
        consultation_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY (patient_id) REFERENCES patients(id),
        FOREIGN KEY (doctor_id) REFERENCES doctors(id)
    )''')
    
    # Predictions table
    c.execute('''CREATE TABLE IF NOT EXISTS predictions (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        patient_id INTEGER,
        symptoms TEXT,
        predicted_disease TEXT,
        confidence REAL,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY (patient_id) REFERENCES patients(id)
    )''')
    
    # Appointments table
    c.execute('''CREATE TABLE IF NOT EXISTS appointments (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        patient_id INTEGER,
        doctor_id INTEGER,
        appointment_date DATE,
        appointment_time TIME,
        status TEXT DEFAULT 'scheduled',
        reason TEXT,
        notes TEXT,
        consultation_type TEXT DEFAULT 'offline',
        prescription TEXT,
        lab_tests TEXT,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY (patient_id) REFERENCES patients(id),
        FOREIGN KEY (doctor_id) REFERENCES doctors(id)
    )''')
    
    # Feedback table
    c.execute('''CREATE TABLE IF NOT EXISTS feedback (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        patient_id INTEGER,
        doctor_id INTEGER,
        consultation_id INTEGER,
        rating INTEGER,
        comment TEXT,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY (patient_id) REFERENCES patients(id),
        FOREIGN KEY (doctor_id) REFERENCES doctors(id),
        FOREIGN KEY (consultation_id) REFERENCES consultations(id)
    )''')
    
    # Medical records table
    c.execute('''CREATE TABLE IF NOT EXISTS medical_records (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        patient_id INTEGER,
        record_type TEXT,
        record_title TEXT,
        record_data TEXT,
        uploaded_by INTEGER,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY (patient_id) REFERENCES patients(id),
        FOREIGN KEY (uploaded_by) REFERENCES users(id)
    )''')
    
    # Doctor availability table
    c.execute('''CREATE TABLE IF NOT EXISTS doctor_availability (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        doctor_id INTEGER,
        day_of_week TEXT,
        start_time TIME,
        end_time TIME,
        is_available BOOLEAN DEFAULT 1,
        FOREIGN KEY (doctor_id) REFERENCES doctors(id)
    )''')
    
    # Notifications table
    c.execute('''CREATE TABLE IF NOT EXISTS notifications (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id INTEGER,
        title TEXT,
        message TEXT,
        type TEXT,
        is_read BOOLEAN DEFAULT 0,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY (user_id) REFERENCES users(id)
    )''')

    # Email settings table (admin-configurable SMTP)
    c.execute('''CREATE TABLE IF NOT EXISTS email_settings (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        mail_server TEXT DEFAULT 'smtp.gmail.com',
        mail_port INTEGER DEFAULT 587,
        mail_use_tls INTEGER DEFAULT 1,
        mail_username TEXT DEFAULT '',
        mail_password TEXT DEFAULT '',
        mail_sender_name TEXT DEFAULT 'Medical System',
        updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    )''')
    # Insert default row if not exists
    c.execute('INSERT OR IGNORE INTO email_settings (id) VALUES (1)')

    conn.commit()
    conn.close()

def get_db():
    conn = sqlite3.connect('medical.db')
    conn.row_factory = sqlite3.Row
    return conn

# Login required decorator with session timeout
def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            flash('Please login first', 'warning')
            return redirect(url_for('signin_as'))
        
        # Check session timeout (30 minutes of inactivity)
        if 'last_activity' in session:
            last_activity = datetime.fromisoformat(session['last_activity'])
            if datetime.now() - last_activity > timedelta(minutes=30):
                session.clear()
                flash('Session expired due to inactivity. Please login again.', 'warning')
                return redirect(url_for('signin_as'))
        
        # Update last activity time
        session['last_activity'] = datetime.now().isoformat()
        session.permanent = True
        
        return f(*args, **kwargs)
    return decorated_function

# Role required decorator
def role_required(role):
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if 'role' not in session or session['role'] != role:
                flash('Access denied', 'danger')
                return redirect(url_for('signin_as'))
            return f(*args, **kwargs)
        return decorated_function
    return decorator

# Home route - Landing Page
@app.route('/')
def index():
    # If user is already logged in, redirect to their dashboard
    if 'user_id' in session:
        if session['role'] == 'patient':
            return redirect(url_for('patient_dashboard'))
        elif session['role'] == 'doctor':
            return redirect(url_for('doctor_dashboard'))
        elif session['role'] == 'admin':
            return redirect(url_for('admin_dashboard'))
    
    # Show modern landing page for new visitors
    return render_template('landing_modern.html')

# Clear session route for testing
@app.route('/clear-session')
def clear_session():
    session.clear()
    flash('Session cleared successfully', 'info')
    return redirect(url_for('index'))

# Test layout route
@app.route('/test-layout')
def test_layout():
    return render_template('test_layout.html')

# Sign-Up As (Role Selection)
@app.route('/signup-as')
def signup_as():
    return render_template('signup_as_modern.html')

# Sign-In As (Role Selection)
@app.route('/signin-as')
def signin_as():
    return render_template('signin_as.html')

# Sign Up as Doctor - CONTROLLED REGISTRATION
# Redirects to new controlled registration system with verification IDs
@app.route('/signup-doctor', methods=['GET', 'POST'])
def signup_doctor():
    # Redirect to controlled registration page
    return redirect(url_for('doctor_registration_page'))

# Sign Up as Patient
@app.route('/signup-patient', methods=['GET', 'POST'])
def signup_patient():
    if request.method == 'POST':
        name = request.form['name']
        email = request.form['email']
        password = request.form['password']
        phone = request.form['phone']
        age = request.form['age']
        gender = request.form['gender']
        address = request.form['address']
        blood_group = request.form.get('blood_group', '')
        
        hashed_password = generate_password_hash(password)
        
        conn = get_db()
        try:
            c = conn.cursor()
            c.execute('INSERT INTO users (name, email, password, role, phone, age, gender, address) VALUES (?, ?, ?, ?, ?, ?, ?, ?)',
                     (name, email, hashed_password, 'patient', phone, age, gender, address))
            user_id = c.lastrowid
            
            c.execute('INSERT INTO patients (user_id, blood_group) VALUES (?, ?)',
                     (user_id, blood_group))
            
            conn.commit()
            flash('Patient registration successful! Please login.', 'success')
            return redirect(url_for('login_patient'))
        except sqlite3.IntegrityError:
            flash('Email already exists', 'danger')
        finally:
            conn.close()
    
    return render_template('signup_patient.html')

# Login as Doctor
@app.route('/login-doctor', methods=['GET', 'POST'])
def login_doctor():
    if request.method == 'POST':
        email = request.form['email']
        password = request.form['password']
        
        conn = get_db()
        user = conn.execute('SELECT * FROM users WHERE email = ? AND role = ?', (email, 'doctor')).fetchone()
        
        if user and check_password_hash(user['password'], password):
            # Check account status
            account_status = user['account_status'] if 'account_status' in user.keys() and user['account_status'] else 'active'
            
            if account_status == 'pending':
                # Get verification ID
                doctor = conn.execute('SELECT verification_id FROM doctors WHERE user_id = ?', (user['id'],)).fetchone()
                verification_id = doctor['verification_id'] if doctor and 'verification_id' in doctor.keys() else 'N/A'
                conn.close()
                
                flash(f'⏳ Your account is pending verification. Verification ID: {verification_id}. You will receive an email once your account is verified by admin.', 'warning')
                return render_template('login_doctor.html', pending_verification=True, verification_id=verification_id)
            
            elif account_status == 'rejected':
                conn.close()
                flash('❌ Your account verification was rejected. Please contact support for more information.', 'danger')
                return render_template('login_doctor.html')
            
            elif account_status == 'suspended':
                conn.close()
                flash('⚠️ Your account has been suspended. Please contact support.', 'danger')
                return render_template('login_doctor.html')
            
            # Account is active - allow login
            conn.close()
            session['user_id'] = user['id']
            session['name'] = user['name']
            session['role'] = user['role']
            session['last_activity'] = datetime.now().isoformat()
            session.permanent = True
            flash(f'Welcome Dr. {user["name"]}!', 'success')
            return redirect(url_for('doctor_dashboard'))
        else:
            conn.close()
            flash('Invalid credentials', 'danger')
    
    return render_template('login_doctor.html')

# Login as Patient
@app.route('/login-patient', methods=['GET', 'POST'])
def login_patient():
    if request.method == 'POST':
        email = request.form['email']
        password = request.form['password']
        
        conn = get_db()
        user = conn.execute('SELECT * FROM users WHERE email = ? AND role = ?', (email, 'patient')).fetchone()
        conn.close()
        
        if user and check_password_hash(user['password'], password):
            session['user_id'] = user['id']
            session['name'] = user['name']
            session['role'] = user['role']
            session['last_activity'] = datetime.now().isoformat()
            session.permanent = True
            flash(f'Welcome {user["name"]}!', 'success')
            return redirect(url_for('patient_dashboard'))
        else:
            flash('Invalid credentials', 'danger')
    
    return render_template('login_patient.html')

# Logout
@app.route('/logout')
def logout():
    session.clear()
    flash('Logged out successfully', 'info')
    return redirect(url_for('signin_as'))

# Patient Dashboard with Health Metrics
@app.route('/patient/dashboard')
@app.route('/patient-dashboard')  # Alternative hyphen URL
@login_required
@role_required('patient')
def patient_dashboard():
    conn = get_db()
    patient = conn.execute('SELECT * FROM patients WHERE user_id = ?', (session['user_id'],)).fetchone()
    
    if not patient:
        # Patient record doesn't exist - create it
        conn.execute('INSERT INTO patients (user_id) VALUES (?)', (session['user_id'],))
        conn.commit()
        patient = conn.execute('SELECT * FROM patients WHERE user_id = ?', (session['user_id'],)).fetchone()
    
    if not patient:
        conn.close()
        return render_template('patient_dashboard.html')
    
    # Get health metrics
    total_consultations = conn.execute('SELECT COUNT(*) as count FROM consultations WHERE patient_id = ?', 
                                      (patient['id'],)).fetchone()['count']
    
    pending_consultations = conn.execute('SELECT COUNT(*) as count FROM consultations WHERE patient_id = ? AND status = "pending"', 
                                        (patient['id'],)).fetchone()['count']
    
    completed_consultations = conn.execute('SELECT COUNT(*) as count FROM consultations WHERE patient_id = ? AND status = "completed"', 
                                          (patient['id'],)).fetchone()['count']
    
    # Upcoming appointments
    upcoming_appointments = conn.execute('''SELECT a.*, d.specialization, u.name as doctor_name
                                           FROM appointments a
                                           JOIN doctors d ON a.doctor_id = d.id
                                           JOIN users u ON d.user_id = u.id
                                           WHERE a.patient_id = ? 
                                           AND a.status IN ('pending', 'scheduled', 'confirmed', 'rescheduled', 'emergency')
                                           AND DATE(a.appointment_date) >= DATE('now')
                                           ORDER BY a.appointment_date ASC, a.appointment_time ASC
                                           LIMIT 3''', (patient['id'],)).fetchall()
    
    # Recent consultations (avoid duplicates)
    recent_consultations = conn.execute('''SELECT DISTINCT c.id, c.consultation_date, c.status, c.prescription,
                                          u.name as doctor_name, d.specialization
                                          FROM consultations c
                                          JOIN doctors d ON c.doctor_id = d.id
                                          JOIN users u ON d.user_id = u.id
                                          WHERE c.patient_id = ?
                                          GROUP BY c.doctor_id, DATE(c.consultation_date)
                                          ORDER BY c.consultation_date DESC
                                          LIMIT 3''', (patient['id'],)).fetchall()
    
    # Recent predictions
    recent_predictions = conn.execute('''SELECT * FROM predictions 
                                        WHERE patient_id = ?
                                        ORDER BY created_at DESC
                                        LIMIT 3''', (patient['id'],)).fetchall()
    
    # Medical records count
    medical_records_count = conn.execute('SELECT COUNT(*) as count FROM medical_records WHERE patient_id = ?', 
                                        (patient['id'],)).fetchone()['count']
    
    # Format data for modern template
    upcoming_count = len(upcoming_appointments) if upcoming_appointments else 0
    prescriptions_count = conn.execute('''SELECT COUNT(*) as count FROM appointments 
                                         WHERE patient_id = ? AND prescription IS NOT NULL''', 
                                       (patient['id'],)).fetchone()['count']
    
    conn.close()
    
    health_stats = {
        'total_consultations': total_consultations,
        'pending_consultations': pending_consultations,
        'completed_consultations': completed_consultations,
        'medical_records_count': medical_records_count,
        'upcoming_appointments': upcoming_appointments,
        'recent_consultations': recent_consultations,
        'recent_predictions': recent_predictions
    }
    
    # Format appointments for display
    formatted_appointments = []
    for apt in upcoming_appointments:
        # Handle both 'mode' and 'consultation_type' columns for backward compatibility
        appointment_mode = apt['mode'] if 'mode' in apt.keys() else (apt['consultation_type'] if 'consultation_type' in apt.keys() else 'online')
        formatted_appointments.append({
            'id': apt['id'],
            'doctor_name': apt['doctor_name'],
            'specialization': apt['specialization'],
            'date': apt['appointment_date'],
            'time': apt['appointment_time'],
            'mode': appointment_mode,
            'status': apt['status'],
            'reason': apt['reason'] if 'reason' in apt.keys() and apt['reason'] else ''
        })
    
    # Format consultations for display
    formatted_consultations = []
    for cons in recent_consultations:
        formatted_consultations.append({
            'id': cons['id'],
            'doctor_name': cons['doctor_name'],
            'specialization': cons['specialization'],
            'date': cons['consultation_date'],
            'prescription': cons['prescription'] if 'prescription' in cons.keys() and cons['prescription'] else ''
        })
    
    return render_template('patient_dashboard.html', 
                         health_stats=health_stats,
                         upcoming_count=upcoming_count,
                         total_consultations=total_consultations,
                         prescriptions_count=prescriptions_count,
                         records_count=medical_records_count,
                         upcoming_appointments=formatted_appointments,
                         recent_consultations=formatted_consultations)

# Patient Profile
@app.route('/patient/profile')
@login_required
@role_required('patient')
def patient_profile():
    conn = get_db()
    user = conn.execute('SELECT * FROM users WHERE id = ?', (session['user_id'],)).fetchone()
    patient = conn.execute('SELECT * FROM patients WHERE user_id = ?', (session['user_id'],)).fetchone()
    conn.close()
    return render_template('patient_profile.html', user=user, patient=patient)

# Doctor Profile
@app.route('/doctor/profile')
@login_required
@role_required('doctor')
def doctor_profile():
    conn = get_db()
    user = conn.execute('SELECT * FROM users WHERE id = ?', (session['user_id'],)).fetchone()
    doctor = conn.execute('SELECT * FROM doctors WHERE user_id = ?', (session['user_id'],)).fetchone()
    conn.close()
    return render_template('doctor_profile.html', user=user, doctor=doctor)

# Consult a Doctor (List of doctors) with Advanced Search
@app.route('/patient/consult-doctor')
@login_required
@role_required('patient')
def consult_doctor():
    # Get search parameters
    search = request.args.get('search', '')
    specialization = request.args.get('specialization', 'all')
    min_rating = request.args.get('min_rating', '0')
    sort_by = request.args.get('sort_by', 'name')
    
    conn = get_db()
    
    # Build query with filters
    query = '''SELECT d.id as doctor_id, d.*, u.name, u.email, u.phone,
               COALESCE(AVG(f.rating), 0) as avg_rating,
               COUNT(DISTINCT c.id) as total_consultations
               FROM doctors d
               JOIN users u ON u.id = d.user_id 
               LEFT JOIN feedback f ON f.doctor_id = d.id
               LEFT JOIN consultations c ON c.doctor_id = d.id
               WHERE u.role = "doctor" AND d.available = 1'''
    
    params = []
    
    # Search filter
    if search:
        query += ' AND (u.name LIKE ? OR d.specialization LIKE ?)'
        params.extend([f'%{search}%', f'%{search}%'])
    
    # Specialization filter
    if specialization != 'all':
        query += ' AND d.specialization = ?'
        params.append(specialization)
    
    query += ' GROUP BY d.id'
    
    # Rating filter
    if min_rating != '0':
        query += ' HAVING avg_rating >= ?'
        params.append(float(min_rating))
    
    # Sorting
    if sort_by == 'rating':
        query += ' ORDER BY avg_rating DESC'
    elif sort_by == 'experience':
        query += ' ORDER BY d.experience DESC'
    elif sort_by == 'fee_low':
        query += ' ORDER BY d.consultation_fee ASC'
    elif sort_by == 'fee_high':
        query += ' ORDER BY d.consultation_fee DESC'
    else:
        query += ' ORDER BY u.name ASC'
    
    doctors = conn.execute(query, params).fetchall()
    
    # Get all specializations for filter dropdown
    specializations = conn.execute('''SELECT DISTINCT specialization 
                                     FROM doctors 
                                     WHERE specialization IS NOT NULL 
                                     ORDER BY specialization''').fetchall()
    
    conn.close()
    
    return render_template('consult_doctor.html', 
                         doctors=doctors, 
                         specializations=specializations,
                         search=search,
                         selected_specialization=specialization,
                         min_rating=min_rating,
                         sort_by=sort_by)

# Consultation Page
@app.route('/patient/consultation/<int:doctor_id>', methods=['GET', 'POST'])
@login_required
@role_required('patient')
def consultation(doctor_id):
    if request.method == 'POST':
        symptoms = request.form['symptoms']
        
        conn = get_db()
        patient = conn.execute('SELECT * FROM patients WHERE user_id = ?', (session['user_id'],)).fetchone()
        
        # Get the latest prediction for this patient to include predicted disease
        latest_prediction = conn.execute('''
            SELECT predicted_disease, confidence 
            FROM predictions 
            WHERE patient_id = ? 
            ORDER BY created_at DESC 
            LIMIT 1
        ''', (patient['id'],)).fetchone()
        
        # If no recent prediction, try to predict based on current symptoms
        predicted_disease = None
        confidence = None
        
        if latest_prediction:
            predicted_disease = latest_prediction['predicted_disease']
            confidence = latest_prediction['confidence']
        else:
            # Generate prediction for current symptoms
            try:
                prediction_result = predict_disease(symptoms)
                predicted_disease = prediction_result['disease']
                confidence = prediction_result['confidence']
                
                # Save this prediction for future reference
                conn.execute('INSERT INTO predictions (patient_id, symptoms, predicted_disease, confidence) VALUES (?, ?, ?, ?)',
                           (patient['id'], symptoms, predicted_disease, confidence))
            except Exception as e:
                print(f"Error predicting disease: {e}")
                predicted_disease = "General Consultation"
                confidence = 0.0
        
        # Create consultation with predicted disease
        c = conn.cursor()
        c.execute('''INSERT INTO consultations (patient_id, doctor_id, symptoms, predicted_disease, status) 
                     VALUES (?, ?, ?, ?, ?)''',
                 (patient['id'], doctor_id, symptoms, predicted_disease, 'pending'))
        conn.commit()
        conn.close()
        
        flash('Consultation request sent successfully!', 'success')
        return redirect(url_for('consultation_history'))
    
    conn = get_db()
    doctor_user = conn.execute('''SELECT d.id as doctor_id, d.*, u.name, u.email, u.phone
                                  FROM doctors d
                                  JOIN users u ON u.id = d.user_id 
                                  WHERE d.id = ?''', (doctor_id,)).fetchone()
    conn.close()
    
    if not doctor_user:
        flash('Doctor not found', 'danger')
        return redirect(url_for('consult_doctor'))
    
    return render_template('consultation.html', doctor=doctor_user)

@app.route('/patient/consultation-history')
@login_required
@role_required('patient')
def consultation_history():
    conn = get_db()
    patient = conn.execute('SELECT * FROM patients WHERE user_id = ?', (session['user_id'],)).fetchone()
    
    if not patient:
        flash('Patient profile not found. Please contact support.', 'danger')
        conn.close()
        return redirect(url_for('patient_dashboard'))
    
    # Unified consultation history: text consultations + video appointments
    consultations = conn.execute('''
        SELECT 
            c.id,
            c.patient_id,
            c.doctor_id,
            c.symptoms as reason,
            c.diagnosis,
            c.prescription,
            c.predicted_disease,
            c.status,
            c.consultation_date,
            'text' as consultation_type,
            'online' as mode,
            u.name as doctor_name,
            u.email as doctor_email,
            u.phone as doctor_phone,
            d.specialization
        FROM consultations c 
        JOIN doctors d ON c.doctor_id = d.id 
        JOIN users u ON d.user_id = u.id 
        WHERE c.patient_id = ?
        
        UNION ALL
        
        SELECT 
            a.id,
            a.patient_id,
            a.doctor_id,
            a.reason,
            a.notes as diagnosis,
            a.prescription,
            'From Appointment' as predicted_disease,
            a.status,
            datetime(a.appointment_date || ' ' || a.appointment_time) as consultation_date,
            'appointment' as consultation_type,
            a.mode,
            u.name as doctor_name,
            u.email as doctor_email,
            u.phone as doctor_phone,
            d.specialization
        FROM appointments a
        JOIN doctors d ON a.doctor_id = d.id 
        JOIN users u ON d.user_id = u.id 
        WHERE a.patient_id = ?
        AND a.status != 'completed'  -- Exclude completed appointments as they have consultation records
        
        ORDER BY consultation_date DESC
    ''', (patient['id'], patient['id'])).fetchall()
    conn.close()
    return render_template('consultation_history.html', consultations=consultations)

# Doctor Consultation History
@app.route('/doctor/consultations', methods=['GET', 'POST'])
@login_required
@role_required('doctor')
def doctor_consultations():
    conn = get_db()
    doctor = conn.execute('SELECT * FROM doctors WHERE user_id = ?', (session['user_id'],)).fetchone()
    
    # Handle consultation update
    if request.method == 'POST':
        consultation_id = request.form.get('consultation_id')
        diagnosis = request.form.get('diagnosis')
        prescription = request.form.get('prescription')
        status = request.form.get('status')
        
        conn.execute('''UPDATE consultations 
                       SET diagnosis = ?, prescription = ?, status = ? 
                       WHERE id = ?''',
                    (diagnosis, prescription, status, consultation_id))
        
        # ==================== AUTO-TRIGGER NEARBY SERVICES ====================
        # When doctor completes consultation with prescription, notify patient about nearby services
        if status == 'completed' and prescription:
            # Get patient details
            consultation = conn.execute('''
                SELECT c.*, p.user_id as patient_user_id, u.name as patient_name
                FROM consultations c
                JOIN patients p ON c.patient_id = p.id
                JOIN users u ON p.user_id = u.id
                WHERE c.id = ?
            ''', (consultation_id,)).fetchone()
            
            if consultation:
                # Analyze prescription to determine service type
                service_type = analyze_prescription_type(prescription)
                
                # Create smart notification based on prescription content
                if service_type == 'pharmacy':
                    notification_title = '💊 Medicines Prescribed - Find Nearby Stores'
                    notification_message = f'Dr. {session["name"]} has prescribed medicines. Click to find nearby medical stores and pharmacies.'
                elif service_type == 'diagnostic':
                    notification_title = '🔬 Tests Recommended - Find Nearby Labs'
                    notification_message = f'Dr. {session["name"]} has recommended diagnostic tests. Click to find nearby labs and imaging centers.'
                else:  # both
                    notification_title = '🏥 Prescription Ready - Find Nearby Services'
                    notification_message = f'Dr. {session["name"]} has prescribed medicines and tests. Click to find nearby medical stores and diagnostic centers.'
                
                # Insert notification with link to nearby services
                conn.execute('''
                    INSERT INTO notifications (user_id, title, message, type)
                    VALUES (?, ?, ?, ?)
                ''', (consultation['patient_user_id'], 
                      notification_title, 
                      notification_message, 
                      'prescription'))
        
        conn.commit()
        conn.close()
        flash('Consultation updated successfully! Patient has been notified about nearby services.', 'success')
        return redirect(url_for('doctor_consultations'))
    
    # Unified consultation history: text consultations + video appointments
    consultations = conn.execute('''
        SELECT 
            c.id,
            c.patient_id,
            c.doctor_id,
            c.symptoms as reason,
            c.diagnosis,
            c.prescription,
            c.predicted_disease,
            c.status,
            c.consultation_date,
            'text' as consultation_type,
            'online' as mode,
            u.name as patient_name,
            u.email as patient_email,
            u.phone as patient_phone
        FROM consultations c 
        JOIN patients p ON c.patient_id = p.id 
        JOIN users u ON p.user_id = u.id 
        WHERE c.doctor_id = ?
        
        UNION ALL
        
        SELECT 
            a.id,
            a.patient_id,
            a.doctor_id,
            a.reason,
            a.notes as diagnosis,
            a.prescription,
            'From Appointment' as predicted_disease,
            a.status,
            datetime(a.appointment_date || ' ' || a.appointment_time) as consultation_date,
            'appointment' as consultation_type,
            a.mode,
            u.name as patient_name,
            u.email as patient_email,
            u.phone as patient_phone
        FROM appointments a
        JOIN patients p ON a.patient_id = p.id 
        JOIN users u ON p.user_id = u.id 
        WHERE a.doctor_id = ?
        AND a.status != 'completed'  -- Exclude completed appointments as they have consultation records
        
        ORDER BY consultation_date DESC
    ''', (doctor['id'], doctor['id'])).fetchall()
    conn.close()
    return render_template('doctor_consultations.html', consultations=consultations)

# Possible Conditions (Disease Prediction)
@app.route('/patient/possible-conditions', methods=['GET', 'POST'])
@login_required
@role_required('patient')
def possible_conditions():
    if request.method == 'POST':
        symptoms = request.form['symptoms']
        
        # Simple disease prediction logic
        predicted_disease = predict_disease(symptoms)
        
        conn = get_db()
        patient = conn.execute('SELECT * FROM patients WHERE user_id = ?', (session['user_id'],)).fetchone()
        
        c = conn.cursor()
        c.execute('INSERT INTO predictions (patient_id, symptoms, predicted_disease, confidence) VALUES (?, ?, ?, ?)',
                 (patient['id'], symptoms, predicted_disease['disease'], predicted_disease['confidence']))
        conn.commit()
        conn.close()
        
        return render_template('prediction_result.html', prediction=predicted_disease)
    
    return render_template('possible_conditions.html')

def predict_disease(symptoms):
    """ML-based disease prediction using trained model"""
    try:
        import pickle
        
        # Load ML model, vectorizer, and label encoder
        with open('model.pkl', 'rb') as f:
            model = pickle.load(f)
        with open('vectorizer.pkl', 'rb') as f:
            vectorizer = pickle.load(f)
        with open('label_encoder.pkl', 'rb') as f:
            label_encoder = pickle.load(f)
        with open('disease_info.pkl', 'rb') as f:
            disease_info = pickle.load(f)
        
        # Vectorize input symptoms
        symptoms_vector = vectorizer.transform([symptoms.lower()])
        
        # Predict disease
        prediction = model.predict(symptoms_vector)
        predicted_disease = label_encoder.inverse_transform(prediction)[0]
        
        # Get confidence score
        confidence = model.predict_proba(symptoms_vector).max() * 100
        
        # Get disease information
        info = disease_info.get(predicted_disease, {
            'description': 'Please consult a doctor for proper diagnosis.',
            'precautions': ['Consult a healthcare professional'],
            'medications': ['As prescribed by doctor'],
            'diet': ['Maintain a balanced diet'],
            'workout': ['Light exercise as tolerated']
        })
        
        return {
            'disease': predicted_disease,
            'confidence': round(confidence, 2),
            'symptoms': symptoms,
            'description': info.get('description', ''),
            'precautions': info.get('precautions', []),
            'medications': info.get('medications', []),
            'diet': info.get('diet', []),
            'workout': info.get('workout', [])
        }
    
    except Exception as e:
        # Fallback to simple prediction if ML model fails
        print(f"ML Model Error: {e}")
        return {
            'disease': 'Unable to predict',
            'confidence': 0,
            'symptoms': symptoms,
            'description': 'Please consult a doctor for proper diagnosis.',
            'precautions': ['Consult a healthcare professional'],
            'medications': ['As prescribed by doctor'],
            'diet': ['Maintain a balanced diet'],
            'workout': ['Light exercise as tolerated']
        }

# ==================== ADMIN ROUTES ====================

# Admin Login
@app.route('/login-admin', methods=['GET', 'POST'])
def login_admin():
    if request.method == 'POST':
        email = request.form['email']
        password = request.form['password']
        
        conn = get_db()
        user = conn.execute('SELECT * FROM users WHERE email = ? AND role = ?', (email, 'admin')).fetchone()
        conn.close()
        
        if user and check_password_hash(user['password'], password):
            session['user_id'] = user['id']
            session['name'] = user['name']
            session['role'] = user['role']
            session['last_activity'] = datetime.now().isoformat()
            session.permanent = True
            flash(f'Welcome Admin {user["name"]}!', 'success')
            return redirect(url_for('admin_dashboard'))
        else:
            flash('Invalid credentials', 'danger')
    
    return render_template('login_admin.html')

# Admin Dashboard
@app.route('/admin/dashboard')
@app.route('/admin-dashboard')  # Alternative hyphen URL
@login_required
@role_required('admin')
def admin_dashboard():
    conn = get_db()
    
    # Get statistics
    total_users = conn.execute('SELECT COUNT(*) as count FROM users').fetchone()['count']
    total_patients = conn.execute('SELECT COUNT(*) as count FROM users WHERE role = "patient"').fetchone()['count']
    total_doctors = conn.execute('SELECT COUNT(*) as count FROM users WHERE role = "doctor"').fetchone()['count']
    total_consultations = conn.execute('SELECT COUNT(*) as count FROM consultations').fetchone()['count']
    pending_consultations = conn.execute('SELECT COUNT(*) as count FROM consultations WHERE status = "pending"').fetchone()['count']
    completed_consultations = conn.execute('SELECT COUNT(*) as count FROM consultations WHERE status = "completed"').fetchone()['count']
    
    # Recent registrations
    recent_users = conn.execute('''SELECT name, email, role, created_at 
                                   FROM users 
                                   ORDER BY created_at DESC 
                                   LIMIT 10''').fetchall()
    
    # Recent consultations
    recent_consultations = conn.execute('''SELECT c.*, 
                                          u1.name as patient_name,
                                          u2.name as doctor_name
                                          FROM consultations c
                                          JOIN patients p ON c.patient_id = p.id
                                          JOIN users u1 ON p.user_id = u1.id
                                          JOIN doctors d ON c.doctor_id = d.id
                                          JOIN users u2 ON d.user_id = u2.id
                                          ORDER BY c.consultation_date DESC
                                          LIMIT 10''').fetchall()
    
    conn.close()
    
    stats = {
        'total_users': total_users,
        'total_patients': total_patients,
        'total_doctors': total_doctors,
        'total_consultations': total_consultations,
        'pending_consultations': pending_consultations,
        'completed_consultations': completed_consultations
    }
    
    return render_template('admin_dashboard_modern.html', stats=stats, recent_users=recent_users, recent_consultations=recent_consultations)

# Admin - Manage Users
@app.route('/admin/users')
@login_required
@role_required('admin')
def admin_users():
    search = request.args.get('search', '')
    role_filter = request.args.get('role', 'all')
    
    conn = get_db()
    
    # Build query based on filters
    query = '''SELECT u.*, 
               CASE 
                   WHEN u.role = "doctor" THEN d.specialization
                   ELSE NULL
               END as specialization,
               CASE 
                   WHEN u.role = "doctor" THEN d.consultation_fee
                   ELSE NULL
               END as consultation_fee
               FROM users u
               LEFT JOIN doctors d ON u.id = d.user_id
               WHERE 1=1'''
    
    params = []
    
    if search:
        query += ' AND (u.name LIKE ? OR u.email LIKE ?)'
        params.extend([f'%{search}%', f'%{search}%'])
    
    if role_filter != 'all':
        query += ' AND u.role = ?'
        params.append(role_filter)
    
    query += ' ORDER BY u.created_at DESC'
    
    users = conn.execute(query, params).fetchall()
    
    # Get counts for statistics
    total_users = conn.execute('SELECT COUNT(*) as count FROM users').fetchone()['count']
    total_patients = conn.execute('SELECT COUNT(*) as count FROM users WHERE role = "patient"').fetchone()['count']
    total_doctors = conn.execute('SELECT COUNT(*) as count FROM users WHERE role = "doctor"').fetchone()['count']
    
    conn.close()
    
    stats = {
        'total_users': total_users,
        'total_patients': total_patients,
        'total_doctors': total_doctors
    }
    
    return render_template('admin_users.html', users=users, stats=stats, search=search, role_filter=role_filter)

# Admin - Manage Consultations
@app.route('/admin/consultations')
@login_required
@role_required('admin')
def admin_consultations():
    search = request.args.get('search', '')
    status_filter = request.args.get('status', 'all')
    
    conn = get_db()
    
    # Build query based on filters
    query = '''SELECT c.*, 
               u1.name as patient_name, u1.email as patient_email,
               u2.name as doctor_name, u2.email as doctor_email,
               d.specialization
               FROM consultations c
               JOIN patients p ON c.patient_id = p.id
               JOIN users u1 ON p.user_id = u1.id
               JOIN doctors d ON c.doctor_id = d.id
               JOIN users u2 ON d.user_id = u2.id
               WHERE 1=1'''
    
    params = []
    
    if search:
        query += ' AND (u1.name LIKE ? OR u2.name LIKE ? OR c.symptoms LIKE ?)'
        params.extend([f'%{search}%', f'%{search}%', f'%{search}%'])
    
    if status_filter != 'all':
        query += ' AND c.status = ?'
        params.append(status_filter)
    
    query += ' ORDER BY c.consultation_date DESC'
    
    consultations = conn.execute(query, params).fetchall()
    conn.close()
    
    return render_template('admin_consultations.html', consultations=consultations, search=search, status_filter=status_filter)

# Export Users to CSV
@app.route('/admin/export/users')
@login_required
@role_required('admin')
def export_users():
    conn = get_db()
    users = conn.execute('''SELECT u.*, 
                           CASE WHEN u.role = "doctor" THEN d.specialization ELSE NULL END as specialization,
                           CASE WHEN u.role = "doctor" THEN d.consultation_fee ELSE NULL END as consultation_fee
                           FROM users u
                           LEFT JOIN doctors d ON u.id = d.user_id
                           ORDER BY u.created_at DESC''').fetchall()
    conn.close()
    
    # Create CSV
    si = StringIO()
    writer = csv.writer(si)
    writer.writerow(['ID', 'Name', 'Email', 'Role', 'Phone', 'Age', 'Gender', 'Specialization', 'Fee', 'Created At'])
    
    for user in users:
        writer.writerow([
            user['id'],
            user['name'],
            user['email'],
            user['role'],
            user['phone'] or '',
            user['age'] or '',
            user['gender'] or '',
            user['specialization'] or '',
            user['consultation_fee'] or '',
            user['created_at']
        ])
    
    output = make_response(si.getvalue())
    output.headers["Content-Disposition"] = f"attachment; filename=users_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    output.headers["Content-type"] = "text/csv"
    return output

# Export Consultations to CSV
@app.route('/admin/export/consultations')
@login_required
@role_required('admin')
def export_consultations():
    conn = get_db()
    consultations = conn.execute('''SELECT c.*, 
                                    u1.name as patient_name, u1.email as patient_email,
                                    u2.name as doctor_name, u2.email as doctor_email,
                                    d.specialization
                                    FROM consultations c
                                    JOIN patients p ON c.patient_id = p.id
                                    JOIN users u1 ON p.user_id = u1.id
                                    JOIN doctors d ON c.doctor_id = d.id
                                    JOIN users u2 ON d.user_id = u2.id
                                    ORDER BY c.consultation_date DESC''').fetchall()
    conn.close()
    
    # Create CSV
    si = StringIO()
    writer = csv.writer(si)
    writer.writerow(['ID', 'Patient', 'Patient Email', 'Doctor', 'Doctor Email', 'Specialization', 'Symptoms', 'Diagnosis', 'Prescription', 'Status', 'Date'])
    
    for c in consultations:
        writer.writerow([
            c['id'],
            c['patient_name'],
            c['patient_email'],
            c['doctor_name'],
            c['doctor_email'],
            c['specialization'],
            c['symptoms'],
            c['diagnosis'] or '',
            c['prescription'] or '',
            c['status'],
            c['consultation_date']
        ])
    
    output = make_response(si.getvalue())
    output.headers["Content-Disposition"] = f"attachment; filename=consultations_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    output.headers["Content-type"] = "text/csv"
    return output

# Export Users to Excel
@app.route('/admin/export/users-excel')
@login_required
@role_required('admin')
def export_users_excel():
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        
        conn = get_db()
        users = conn.execute('''SELECT u.*, 
                               CASE WHEN u.role = "doctor" THEN d.specialization ELSE NULL END as specialization,
                               CASE WHEN u.role = "doctor" THEN d.consultation_fee ELSE NULL END as consultation_fee
                               FROM users u
                               LEFT JOIN doctors d ON u.id = d.user_id
                               ORDER BY u.created_at DESC''').fetchall()
        conn.close()
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Users"
        
        # Headers with styling
        headers = ['ID', 'Name', 'Email', 'Role', 'Phone', 'Age', 'Gender', 'Specialization', 'Fee', 'Created At']
        ws.append(headers)
        
        # Style headers
        header_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Add data
        for user in users:
            ws.append([
                user['id'],
                user['name'],
                user['email'],
                user['role'],
                user['phone'] or '',
                user['age'] or '',
                user['gender'] or '',
                user['specialization'] or '',
                user['consultation_fee'] or '',
                user['created_at']
            ])
        
        # Adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        from io import BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = make_response(output.read())
        response.headers["Content-Disposition"] = f"attachment; filename=users_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response.headers["Content-type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        return response
        
    except ImportError:
        flash('Excel export requires openpyxl. Install it with: pip install openpyxl', 'warning')
        return redirect(url_for('admin_users'))

# Export Consultations to Excel
@app.route('/admin/export/consultations-excel')
@login_required
@role_required('admin')
def export_consultations_excel():
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        
        conn = get_db()
        consultations = conn.execute('''SELECT c.*, 
                                        u1.name as patient_name, u1.email as patient_email,
                                        u2.name as doctor_name, u2.email as doctor_email,
                                        d.specialization
                                        FROM consultations c
                                        JOIN patients p ON c.patient_id = p.id
                                        JOIN users u1 ON p.user_id = u1.id
                                        JOIN doctors d ON c.doctor_id = d.id
                                        JOIN users u2 ON d.user_id = u2.id
                                        ORDER BY c.consultation_date DESC''').fetchall()
        conn.close()
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Consultations"
        
        # Headers
        headers = ['ID', 'Patient', 'Patient Email', 'Doctor', 'Doctor Email', 'Specialization', 
                  'Symptoms', 'Diagnosis', 'Prescription', 'Status', 'Date']
        ws.append(headers)
        
        # Style headers
        header_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Add data
        for c in consultations:
            ws.append([
                c['id'],
                c['patient_name'],
                c['patient_email'],
                c['doctor_name'],
                c['doctor_email'],
                c['specialization'],
                c['symptoms'],
                c['diagnosis'] or '',
                c['prescription'] or '',
                c['status'],
                c['consultation_date']
            ])
        
        # Adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        from io import BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = make_response(output.read())
        response.headers["Content-Disposition"] = f"attachment; filename=consultations_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response.headers["Content-type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        return response
        
    except ImportError:
        flash('Excel export requires openpyxl. Install it with: pip install openpyxl', 'warning')
        return redirect(url_for('admin_consultations'))

# Admin Reports Page
@app.route('/admin/reports')
@login_required
@role_required('admin')
def admin_reports():
    conn = get_db()
    
    # Get comprehensive statistics
    total_users = conn.execute('SELECT COUNT(*) as count FROM users').fetchone()['count']
    total_patients = conn.execute('SELECT COUNT(*) as count FROM users WHERE role = "patient"').fetchone()['count']
    total_doctors = conn.execute('SELECT COUNT(*) as count FROM users WHERE role = "doctor"').fetchone()['count']
    total_consultations = conn.execute('SELECT COUNT(*) as count FROM consultations').fetchone()['count']
    pending_consultations = conn.execute('SELECT COUNT(*) as count FROM consultations WHERE status = "pending"').fetchone()['count']
    completed_consultations = conn.execute('SELECT COUNT(*) as count FROM consultations WHERE status = "completed"').fetchone()['count']
    
    # Get consultations by month (last 6 months)
    monthly_consultations = conn.execute('''
        SELECT strftime('%Y-%m', consultation_date) as month, COUNT(*) as count
        FROM consultations
        WHERE consultation_date >= date('now', '-6 months')
        GROUP BY month
        ORDER BY month
    ''').fetchall()
    
    # Get top doctors by consultation count
    top_doctors = conn.execute('''
        SELECT u.name, d.specialization, COUNT(c.id) as consultation_count
        FROM doctors d
        JOIN users u ON d.user_id = u.id
        LEFT JOIN consultations c ON d.id = c.doctor_id
        GROUP BY d.id
        ORDER BY consultation_count DESC
        LIMIT 5
    ''').fetchall()
    
    # Get most common symptoms - improved tracking
    # Use predictions table which has structured symptom data
    all_predictions = conn.execute('SELECT symptoms FROM predictions WHERE symptoms IS NOT NULL').fetchall()
    symptom_counts = {}
    
    # Process predictions (comma-separated symptoms)
    for pred in all_predictions:
        if pred['symptoms']:
            # Split by comma for multi-word symptoms
            symptoms = [s.strip().lower() for s in pred['symptoms'].split(',')]
            for symptom in symptoms:
                if len(symptom) > 2:  # Only count symptoms longer than 2 chars
                    symptom_counts[symptom] = symptom_counts.get(symptom, 0) + 1
    
    # Also process consultation symptoms (fallback)
    all_consultations = conn.execute('SELECT symptoms FROM consultations WHERE symptoms IS NOT NULL').fetchall()
    for consult in all_consultations:
        if consult['symptoms']:
            # Try comma-separated first, then space-separated
            if ',' in consult['symptoms']:
                symptoms = [s.strip().lower() for s in consult['symptoms'].split(',')]
            else:
                symptoms = [s.strip().lower() for s in consult['symptoms'].split()]
            
            for symptom in symptoms:
                symptom = symptom.strip('.,!?')
                if len(symptom) > 2:  # Only count symptoms longer than 2 chars
                    symptom_counts[symptom] = symptom_counts.get(symptom, 0) + 1
    
    top_symptoms = sorted(symptom_counts.items(), key=lambda x: x[1], reverse=True)[:10]
    
    # Get recent activity
    recent_registrations = conn.execute('''
        SELECT name, email, role, created_at
        FROM users
        ORDER BY created_at DESC
        LIMIT 10
    ''').fetchall()
    
    conn.close()
    
    stats = {
        'total_users': total_users,
        'total_patients': total_patients,
        'total_doctors': total_doctors,
        'total_consultations': total_consultations,
        'pending_consultations': pending_consultations,
        'completed_consultations': completed_consultations
    }
    
    return render_template('admin_reports.html', 
                         stats=stats, 
                         monthly_consultations=monthly_consultations,
                         top_doctors=top_doctors,
                         top_symptoms=top_symptoms,
                         recent_registrations=recent_registrations)

# Admin Settings Page
@app.route('/admin/settings', methods=['GET', 'POST'])
@login_required
@role_required('admin')
def admin_settings():
    if request.method == 'POST':
        # Handle settings update
        flash('Settings updated successfully!', 'success')
        return redirect(url_for('admin_settings'))
    
    conn = get_db()
    
    # Get system information
    total_users = conn.execute('SELECT COUNT(*) as count FROM users').fetchone()['count']
    total_consultations = conn.execute('SELECT COUNT(*) as count FROM consultations').fetchone()['count']
    
    # Get database size (approximate)
    import os
    db_size = os.path.getsize('medical.db') / 1024  # Size in KB
    
    conn.close()
    
    system_info = {
        'total_users': total_users,
        'total_consultations': total_consultations,
        'db_size': f'{db_size:.2f} KB',
        'version': '1.0.0',
        'status': 'Active'
    }

    email_cfg = get_email_settings() or {}
    return render_template('admin_settings.html', system_info=system_info, email_cfg=email_cfg)

# Update Settings
@app.route('/admin/update-settings', methods=['POST'])
@login_required
@role_required('admin')
def admin_update_settings():
    # In a real application, these would be saved to a settings table
    flash('Settings updated successfully!', 'success')
    return redirect(url_for('admin_settings'))

# Update Notifications
@app.route('/admin/update-notifications', methods=['POST'])
@login_required
@role_required('admin')
def admin_update_notifications():
    # In a real application, these would be saved to a settings table
    flash('Notification settings updated successfully!', 'success')
    return redirect(url_for('admin_settings'))

# ==================== EMAIL CONFIGURATION (Admin) ====================

@app.route('/admin/save-email-config', methods=['POST'])
@login_required
@role_required('admin')
def admin_save_email_config():
    """Save SMTP email configuration to the database."""
    mail_server      = request.form.get('mail_server', 'smtp.gmail.com').strip()
    mail_port        = int(request.form.get('mail_port', 587))
    mail_use_tls     = 1 if request.form.get('mail_use_tls') == 'on' else 0
    mail_username    = request.form.get('mail_username', '').strip()
    mail_password    = request.form.get('mail_password', '').strip()
    mail_sender_name = request.form.get('mail_sender_name', 'Medical System').strip()

    # Keep existing password if blank (don't overwrite with empty)
    conn = get_db()
    if not mail_password:
        existing = conn.execute('SELECT mail_password FROM email_settings WHERE id = 1').fetchone()
        if existing:
            mail_password = existing['mail_password']

    conn.execute('''UPDATE email_settings SET
        mail_server=?, mail_port=?, mail_use_tls=?,
        mail_username=?, mail_password=?, mail_sender_name=?,
        updated_at=CURRENT_TIMESTAMP
        WHERE id=1''',
        (mail_server, mail_port, mail_use_tls, mail_username, mail_password, mail_sender_name))
    conn.commit()
    conn.close()

    flash('✅ Email configuration saved successfully!', 'success')
    return redirect(url_for('admin_settings'))


@app.route('/admin/test-email', methods=['POST'])
@login_required
@role_required('admin')
def admin_test_email():
    """Send a test email to verify SMTP configuration."""
    test_recipient = request.form.get('test_recipient', '').strip()
    if not test_recipient:
        return jsonify({'success': False, 'message': 'Please enter a recipient email address.'})

    cfg = get_email_settings()
    if not cfg or not cfg['username'] or not cfg['password']:
        return jsonify({'success': False, 'message': 'Email not configured. Please save SMTP settings first.'})

    import smtplib
    from email.mime.text import MIMEText
    from email.mime.multipart import MIMEMultipart

    html = """
<html><body style="font-family:Arial,sans-serif;padding:20px;">
<div style="max-width:500px;margin:auto;background:white;border-radius:12px;padding:28px;box-shadow:0 4px 20px rgba(0,0,0,0.1);">
  <h2 style="color:#667eea;">✅ Email Test Successful!</h2>
  <p>Your SMTP configuration is working correctly.</p>
  <p>Doctors will now receive real-time email notifications whenever a patient books an appointment.</p>
  <hr style="border:none;border-top:1px solid #e5e7eb;margin:20px 0;">
  <p style="color:#9ca3af;font-size:12px;">Medical System – Automated Test</p>
</div>
</body></html>"""

    try:
        msg = MIMEMultipart('alternative')
        msg['Subject'] = '✅ Medical System – Email Test'
        msg['From']    = f"{cfg['sender_name']} <{cfg['username']}>"
        msg['To']      = test_recipient
        msg.attach(MIMEText('Email test successful! Your SMTP configuration is working.', 'plain'))
        msg.attach(MIMEText(html, 'html'))

        if cfg['use_tls']:
            server = smtplib.SMTP(cfg['server'], cfg['port'], timeout=15)
            server.ehlo()
            server.starttls()
        else:
            server = smtplib.SMTP_SSL(cfg['server'], cfg['port'], timeout=15)

        server.login(cfg['username'], cfg['password'])
        server.sendmail(cfg['username'], [test_recipient], msg.as_string())
        server.quit()

        return jsonify({'success': True, 'message': f'Test email sent to {test_recipient}!'})
    except Exception as e:
        return jsonify({'success': False, 'message': f'Failed: {str(e)}'})

# ==================== ADMIN USER MANAGEMENT ====================

# Add User
@app.route('/admin/add-user', methods=['POST'])
@login_required
@role_required('admin')
def admin_add_user():
    name = request.form['name']
    email = request.form['email']
    password = request.form['password']
    role = request.form['role']
    phone = request.form.get('phone', '')
    
    hashed_password = generate_password_hash(password)
    
    conn = get_db()
    try:
        c = conn.cursor()
        c.execute('INSERT INTO users (name, email, password, role, phone) VALUES (?, ?, ?, ?, ?)',
                 (name, email, hashed_password, role, phone))
        user_id = c.lastrowid
        
        # Create corresponding patient or doctor record
        if role == 'patient':
            c.execute('INSERT INTO patients (user_id) VALUES (?)', (user_id,))
        elif role == 'doctor':
            c.execute('INSERT INTO doctors (user_id, specialization, qualification, experience, consultation_fee) VALUES (?, ?, ?, ?, ?)',
                     (user_id, 'General', 'MBBS', 0, 500))
        
        conn.commit()
        flash(f'{role.capitalize()} added successfully!', 'success')
    except sqlite3.IntegrityError:
        flash('Email already exists', 'danger')
    finally:
        conn.close()
    
    return redirect(url_for('admin_users'))

# Edit User
@app.route('/admin/edit-user/<int:user_id>', methods=['POST'])
@login_required
@role_required('admin')
def admin_edit_user(user_id):
    name = request.form['name']
    email = request.form['email']
    phone = request.form.get('phone', '')
    
    conn = get_db()
    try:
        conn.execute('UPDATE users SET name = ?, email = ?, phone = ? WHERE id = ?',
                    (name, email, phone, user_id))
        conn.commit()
        flash('User updated successfully!', 'success')
    except sqlite3.IntegrityError:
        flash('Email already exists', 'danger')
    finally:
        conn.close()
    
    return redirect(url_for('admin_users'))

# Delete User
@app.route('/admin/delete-user/<int:user_id>')
@login_required
@role_required('admin')
def admin_delete_user(user_id):
    conn = get_db()
    
    # Check if user is admin
    user = conn.execute('SELECT role FROM users WHERE id = ?', (user_id,)).fetchone()
    
    if user and user['role'] != 'admin':
        # Delete related records first
        if user['role'] == 'patient':
            patient = conn.execute('SELECT id FROM patients WHERE user_id = ?', (user_id,)).fetchone()
            if patient:
                conn.execute('DELETE FROM consultations WHERE patient_id = ?', (patient['id'],))
                conn.execute('DELETE FROM appointments WHERE patient_id = ?', (patient['id'],))
                conn.execute('DELETE FROM feedback WHERE patient_id = ?', (patient['id'],))
                conn.execute('DELETE FROM medical_records WHERE patient_id = ?', (patient['id'],))
                conn.execute('DELETE FROM patients WHERE user_id = ?', (user_id,))
        elif user['role'] == 'doctor':
            doctor = conn.execute('SELECT id FROM doctors WHERE user_id = ?', (user_id,)).fetchone()
            if doctor:
                conn.execute('DELETE FROM consultations WHERE doctor_id = ?', (doctor['id'],))
                conn.execute('DELETE FROM appointments WHERE doctor_id = ?', (doctor['id'],))
                conn.execute('DELETE FROM feedback WHERE doctor_id = ?', (doctor['id'],))
                conn.execute('DELETE FROM doctor_availability WHERE doctor_id = ?', (doctor['id'],))
                conn.execute('DELETE FROM doctors WHERE user_id = ?', (user_id,))
        
        # Delete notifications
        conn.execute('DELETE FROM notifications WHERE user_id = ?', (user_id,))
        
        # Delete user
        conn.execute('DELETE FROM users WHERE id = ?', (user_id,))
        conn.commit()
        flash('User deleted successfully!', 'success')
    else:
        flash('Cannot delete admin users!', 'danger')
    
    conn.close()
    return redirect(url_for('admin_users'))

# ==================== DOCTOR STATISTICS ====================

# Enhanced Doctor Dashboard with Statistics
@app.route('/doctor/dashboard')
@app.route('/doctor-dashboard')  # Alternative hyphen URL
@login_required
@role_required('doctor')
def doctor_dashboard():
    conn = get_db()
    doctor = conn.execute('SELECT * FROM doctors WHERE user_id = ?', (session['user_id'],)).fetchone()
    
    # Get statistics
    total_consultations = conn.execute('SELECT COUNT(*) as count FROM consultations WHERE doctor_id = ?', (doctor['id'],)).fetchone()['count']
    pending_consultations = conn.execute('SELECT COUNT(*) as count FROM consultations WHERE doctor_id = ? AND status = "pending"', (doctor['id'],)).fetchone()['count']
    completed_consultations = conn.execute('SELECT COUNT(*) as count FROM consultations WHERE doctor_id = ? AND status = "completed"', (doctor['id'],)).fetchone()['count']
    
    # Today's consultations
    today_consultations = conn.execute('''SELECT COUNT(*) as count FROM consultations 
                                         WHERE doctor_id = ? AND DATE(consultation_date) = DATE('now')''', 
                                       (doctor['id'],)).fetchone()['count']
    
    # This week's consultations
    week_consultations = conn.execute('''SELECT COUNT(*) as count FROM consultations 
                                        WHERE doctor_id = ? AND DATE(consultation_date) >= DATE('now', '-7 days')''', 
                                      (doctor['id'],)).fetchone()['count']
    
    # This month's consultations
    month_consultations = conn.execute('''SELECT COUNT(*) as count FROM consultations 
                                         WHERE doctor_id = ? AND DATE(consultation_date) >= DATE('now', 'start of month')''', 
                                       (doctor['id'],)).fetchone()['count']
    
    # Total unique patients
    total_patients = conn.execute('SELECT COUNT(DISTINCT patient_id) as count FROM consultations WHERE doctor_id = ?', (doctor['id'],)).fetchone()['count']
    
    # Pending appointments count (for badge)
    pending_appointments = conn.execute('''SELECT COUNT(*) as count FROM appointments 
                                          WHERE doctor_id = ? AND status IN ('pending', 'scheduled')''', 
                                       (doctor['id'],)).fetchone()['count']
    
    # Today's appointments
    today_appointments_list = conn.execute('''SELECT a.*, u.name as patient_name, u.phone as patient_phone
                                             FROM appointments a
                                             JOIN patients p ON a.patient_id = p.id
                                             JOIN users u ON p.user_id = u.id
                                             WHERE a.doctor_id = ? 
                                             AND DATE(a.appointment_date) = DATE('now')
                                             AND a.status IN ('scheduled', 'confirmed', 'emergency', 'pending')
                                             ORDER BY a.appointment_time ASC''', 
                                           (doctor['id'],)).fetchall()
    
    # Upcoming appointments (next 7 days, excluding today)
    upcoming_appointments_list = conn.execute('''SELECT a.*, u.name as patient_name, u.phone as patient_phone
                                                 FROM appointments a
                                                 JOIN patients p ON a.patient_id = p.id
                                                 JOIN users u ON p.user_id = u.id
                                                 WHERE a.doctor_id = ? 
                                                 AND DATE(a.appointment_date) > DATE('now')
                                                 AND DATE(a.appointment_date) <= DATE('now', '+7 days')
                                                 AND a.status IN ('scheduled', 'confirmed', 'emergency', 'pending')
                                                 ORDER BY a.appointment_date ASC, a.appointment_time ASC
                                                 LIMIT 10''', 
                                               (doctor['id'],)).fetchall()
    
    # Recent consultations
    recent_consultations = conn.execute('''SELECT c.*, u.name as patient_name, u.email as patient_email
                                          FROM consultations c
                                          JOIN patients p ON c.patient_id = p.id
                                          JOIN users u ON p.user_id = u.id
                                          WHERE c.doctor_id = ?
                                          ORDER BY c.consultation_date DESC
                                          LIMIT 5''', (doctor['id'],)).fetchall()
    
    # Get doctor's verification ID and specialization
    verification_id = doctor['verification_id'] if doctor and 'verification_id' in doctor.keys() and doctor['verification_id'] else 'Not Assigned'
    specialization = doctor['specialization'] if doctor and 'specialization' in doctor.keys() and doctor['specialization'] else 'General'
    
    # Store in session for navbar
    session['verification_id'] = verification_id
    session['specialization'] = specialization
    
    conn.close()
    
    # Format today's appointments
    formatted_today_appointments = []
    for apt in today_appointments_list:
        formatted_today_appointments.append({
            'id': apt['id'],
            'patient_name': apt['patient_name'],
            'time': apt['appointment_time'],
            'date': apt['appointment_date'],
            'mode': apt['mode'],
            'reason': apt['reason'] if 'reason' in apt.keys() and apt['reason'] else '',
            'status': apt['status']
        })
    
    # Format upcoming appointments
    formatted_upcoming_appointments = []
    for apt in upcoming_appointments_list:
        formatted_upcoming_appointments.append({
            'id': apt['id'],
            'patient_name': apt['patient_name'],
            'time': apt['appointment_time'],
            'date': apt['appointment_date'],
            'mode': apt['mode'],
            'reason': apt['reason'] if 'reason' in apt.keys() and apt['reason'] else '',
            'status': apt['status']
        })
    
    # Format recent consultations
    formatted_consultations = []
    for cons in recent_consultations:
        formatted_consultations.append({
            'id': cons['id'],
            'patient_name': cons['patient_name'],
            'date': cons['consultation_date'],
            'time': cons['consultation_time'] if 'consultation_time' in cons.keys() and cons['consultation_time'] else '',
            'diagnosis': cons['diagnosis'] if 'diagnosis' in cons.keys() and cons['diagnosis'] else '',
            'prescription': cons['prescription'] if 'prescription' in cons.keys() and cons['prescription'] else ''
        })
    
    stats = {
        'total_consultations': total_consultations,
        'pending_consultations': pending_consultations,
        'completed_consultations': completed_consultations,
        'today_consultations': today_consultations,
        'week_consultations': week_consultations,
        'month_consultations': month_consultations,
        'total_patients': total_patients,
        'pending_appointments': pending_appointments
    }
    
    return render_template('doctor_dashboard_modern.html', 
                         stats=stats, 
                         today_appointments=len(formatted_today_appointments),
                         pending_appointments=pending_appointments,
                         total_patients=total_patients,
                         today_appointments_list=formatted_today_appointments,
                         upcoming_appointments_list=formatted_upcoming_appointments,
                         recent_consultations=formatted_consultations,
                         pending_consultations=pending_consultations,
                         verification_id=verification_id,
                         specialization=specialization)

# ==================== APPOINTMENT BOOKING ====================

@app.route('/patient/book-appointment/<int:doctor_id>', methods=['GET', 'POST'])
@login_required
@role_required('patient')
def book_appointment(doctor_id):
    if request.method == 'POST':
        appointment_date = request.form['appointment_date']
        appointment_time = request.form['appointment_time']
        reason = request.form['reason']
        consultation_mode = request.form.get('consultation_mode', 'online')  # Get consultation mode
        is_emergency = request.form.get('is_emergency', 'off') == 'on'
        
        conn = get_db()
        patient = conn.execute('SELECT * FROM patients WHERE user_id = ?', (session['user_id'],)).fetchone()
        
        if not patient:
            conn.execute('INSERT INTO patients (user_id) VALUES (?)', (session['user_id'],))
            conn.commit()
            patient = conn.execute('SELECT * FROM patients WHERE user_id = ?', (session['user_id'],)).fetchone()
        
        patient_user = conn.execute('SELECT name, email FROM users WHERE id = ?', (session['user_id'],)).fetchone()
        
        # Create appointment - status should be 'pending' so doctor can accept/reject
        # Only emergency appointments bypass doctor approval
        status = 'emergency' if is_emergency else 'pending'
        cursor = conn.cursor()
        cursor.execute('''INSERT INTO appointments (patient_id, doctor_id, appointment_date, appointment_time, reason, status, mode)
                       VALUES (?, ?, ?, ?, ?, ?, ?)''',
                    (patient['id'], doctor_id, appointment_date, appointment_time, reason, status, consultation_mode))
        appointment_id = cursor.lastrowid  # Get the appointment ID
        
        # Get doctor details
        doctor = conn.execute('SELECT d.*, u.name as doctor_name, u.email as doctor_email FROM doctors d JOIN users u ON d.user_id = u.id WHERE d.id = ?', (doctor_id,)).fetchone()
        
        # Send email confirmation to patient
        send_appointment_confirmation(
            patient_email=patient_user['email'],
            patient_name=patient_user['name'],
            doctor_name=doctor['doctor_name'],
            specialization=doctor['specialization'],
            date=appointment_date,
            time=appointment_time,
            mode=consultation_mode,  # Use the selected mode
            appointment_id=appointment_id
        )
        
        # Send email notification to doctor
        send_doctor_appointment_notification(
            doctor_email=doctor['doctor_email'],
            doctor_name=doctor['doctor_name'],
            patient_name=patient_user['name'],
            date=appointment_date,
            time=appointment_time,
            mode=consultation_mode,  # Use the selected mode
            is_emergency=is_emergency
        )
        
        # Create notification for doctor
        mode_text = "Online (Video Call)" if consultation_mode == 'online' else "In-Person (Clinic Visit)"
        if is_emergency:
            notification_title = '🚨 EMERGENCY APPOINTMENT'
            notification_message = f'URGENT: Emergency {mode_text} appointment from {patient_user["name"]} for {appointment_date} at {appointment_time}. Reason: {reason}'
            notification_type = 'emergency'
        else:
            notification_title = 'New Appointment'
            notification_message = f'New {mode_text} appointment scheduled for {appointment_date} at {appointment_time} with {patient_user["name"]}'
            notification_type = 'appointment'
        
        conn.execute('''INSERT INTO notifications (user_id, title, message, type)
                       VALUES (?, ?, ?, ?)''',
                    (doctor['user_id'], notification_title, notification_message, notification_type))
        
        # Create confirmation notification for patient
        patient_notification = f'Your {"emergency " if is_emergency else ""}{mode_text} appointment with Dr. {doctor["doctor_name"]} has been booked for {appointment_date} at {appointment_time}'
        conn.execute('''INSERT INTO notifications (user_id, title, message, type)
                       VALUES (?, ?, ?, ?)''',
                    (session['user_id'], 'Appointment Confirmed', patient_notification, 'appointment'))
        
        conn.commit()
        conn.close()
        
        # ── Real-time: SSE push to doctor's browser ───────────────────────────
        push_notification_to_doctor(doctor['user_id'], {
            'type':              notification_type,
            'title':             notification_title,
            'message':           notification_message,
            'patient_name':      patient_user['name'],
            'appointment_date':  appointment_date,
            'appointment_time':  appointment_time,
            'consultation_type': 'Emergency' if is_emergency else 'In-Person',
            'symptoms':          reason or '',
        })

        # ── Real-time: Socket.IO push (fallback) ──────────────────────────────
        if socketio:
            try:
                socketio.emit('new_appointment', {
                    'title':             notification_title,
                    'message':           notification_message,
                    'type':              notification_type,
                    'patient_name':      patient_user['name'],
                    'appointment_date':  appointment_date,
                    'appointment_time':  appointment_time,
                    'consultation_type': consultation_mode,
                    'symptoms':          reason or ''
                }, room=f"doctor_{doctor['user_id']}", namespace='/')
            except Exception as sio_err:
                print(f"[Socket.IO] {sio_err}")

        if is_emergency:
            flash('🚨 Emergency appointment booked! Doctor has been notified immediately.', 'warning')
            return redirect(url_for('patient_appointments'))
        else:
            return redirect(url_for('payment.checkout', appointment_id=appointment_id))
    
    conn = get_db()
    doctor = conn.execute('''SELECT d.*, u.name, u.email 
                            FROM doctors d 
                            JOIN users u ON d.user_id = u.id 
                            WHERE d.id = ?''', (doctor_id,)).fetchone()
    
    # Get doctor availability
    availability = conn.execute('''SELECT * FROM doctor_availability 
                                  WHERE doctor_id = ? AND is_available = 1''', 
                               (doctor_id,)).fetchall()
    conn.close()
    
    # Get today's date for min date in form
    from datetime import date
    today = date.today().strftime('%Y-%m-%d')
    
    return render_template('book_appointment_modern.html', doctor=doctor, availability=availability, today=today)

@app.route('/patient/appointments')
@login_required
@role_required('patient')
def patient_appointments():
    conn = get_db()
    patient = conn.execute('SELECT * FROM patients WHERE user_id = ?', (session['user_id'],)).fetchone()
    
    if not patient:
        # Patient record doesn't exist - create it
        conn.execute('INSERT INTO patients (user_id) VALUES (?)', (session['user_id'],))
        conn.commit()
        patient = conn.execute('SELECT * FROM patients WHERE user_id = ?', (session['user_id'],)).fetchone()
    
    appointments = conn.execute('''SELECT a.*, d.specialization, u.name as doctor_name
                                  FROM appointments a
                                  JOIN doctors d ON a.doctor_id = d.id
                                  JOIN users u ON d.user_id = u.id
                                  WHERE a.patient_id = ?
                                  ORDER BY a.appointment_date DESC, a.appointment_time DESC''',
                               (patient['id'],)).fetchall()
    conn.close()
    
    return render_template('patient_appointments.html', appointments=appointments)

@app.route('/doctor/appointments')
@login_required
@role_required('doctor')
def doctor_appointments():
    conn = get_db()
    doctor = conn.execute('SELECT * FROM doctors WHERE user_id = ?', (session['user_id'],)).fetchone()
    
    appointments = conn.execute('''SELECT a.*, u.name as patient_name, u.email as patient_email, u.phone as patient_phone
                                  FROM appointments a
                                  JOIN patients p ON a.patient_id = p.id
                                  JOIN users u ON p.user_id = u.id
                                  WHERE a.doctor_id = ?
                                  ORDER BY a.appointment_date DESC, a.appointment_time DESC''',
                               (doctor['id'],)).fetchall()
    conn.close()
    
    return render_template('doctor_appointments.html', appointments=appointments)

# Test route for debugging accept button
@app.route('/test-accept')
@login_required
@role_required('doctor')
def test_accept():
    return '''
    <!DOCTYPE html>
    <html>
    <head>
        <title>Test Accept Button</title>
        <script src="https://cdn.jsdelivr.net/npm/sweetalert2@11"></script>
    </head>
    <body>
        <h1>Test Accept Button - Logged in as Doctor</h1>
        <p>This page tests if the Accept API works</p>
        <button onclick="testAccept()" style="padding: 10px 20px; font-size: 16px;">Test Accept Appointment #4</button>
        
        <div id="output" style="margin-top: 20px; padding: 20px; background: #f0f0f0;">
            <h3>Output:</h3>
            <pre id="log"></pre>
        </div>
        
        <script>
            function log(message) {
                const logEl = document.getElementById('log');
                logEl.textContent += message + '\\n';
                console.log(message);
            }
            
            function testAccept() {
                log('[TEST] Starting accept test...');
                
                Swal.fire({
                    title: 'Accept Appointment?',
                    text: 'This will accept appointment #4',
                    icon: 'question',
                    showCancelButton: true,
                    confirmButtonText: 'Yes, Accept'
                }).then(function(result) {
                    log('[TEST] Swal result: ' + JSON.stringify(result));
                    
                    if (!result.isConfirmed) {
                        log('[TEST] User cancelled');
                        return;
                    }
                    
                    log('[TEST] Sending API request...');
                    
                    fetch('/api/appointment/action', {
                        method: 'POST',
                        headers: { 'Content-Type': 'application/json' },
                        body: JSON.stringify({ appointment_id: 4, action: 'accept' })
                    })
                    .then(function(r) {
                        log('[TEST] Response status: ' + r.status);
                        return r.json();
                    })
                    .then(function(data) {
                        log('[TEST] Response data: ' + JSON.stringify(data));
                        
                        if (data.success) {
                            Swal.fire('Success!', 'Appointment accepted! Check terminal for email logs.', 'success');
                        } else {
                            Swal.fire('Error', data.error, 'error');
                        }
                    })
                    .catch(function(err) {
                        log('[TEST] Error: ' + err.message);
                        Swal.fire('Error', err.message, 'error');
                    });
                });
            }
        </script>
    </body>
    </html>
    '''

# ==================== APPOINTMENT ACCEPT / REJECT API ====================

@app.route('/api/appointment/action', methods=['POST'])
@login_required
@role_required('doctor')
def api_appointment_action():
    """Accept or reject an appointment. Called from doctor_appointments.html."""
    print("[API] Appointment action request received")
    try:
        data           = request.get_json()
        appointment_id = data.get('appointment_id')
        action         = data.get('action')  # 'accept' or 'reject'
        
        print(f"[API] Appointment ID: {appointment_id}, Action: {action}")

        if not appointment_id or action not in ('accept', 'reject'):
            print("[API] Invalid request - missing appointment_id or invalid action")
            return jsonify({'success': False, 'error': 'Invalid request'})

        new_status = 'confirmed' if action == 'accept' else 'cancelled'

        conn = get_db()
        appt = conn.execute('''
            SELECT a.*, u.name as patient_name, u.id as patient_user_id, u.email as patient_email,
                   u2.name as doctor_name
            FROM appointments a
            JOIN patients p ON a.patient_id = p.id
            JOIN users u ON p.user_id = u.id
            JOIN doctors d ON a.doctor_id = d.id
            JOIN users u2 ON d.user_id = u2.id
            WHERE a.id = ?
        ''', (appointment_id,)).fetchone()

        if not appt:
            print(f"[API] Appointment #{appointment_id} not found")
            conn.close()
            return jsonify({'success': False, 'error': 'Appointment not found'})

        print(f"[API] Updating appointment #{appointment_id} status to: {new_status}")
        conn.execute('UPDATE appointments SET status = ? WHERE id = ?', (new_status, appointment_id))

        # Notify patient
        if action == 'accept':
            msg = (f'Your appointment with Dr. {appt["doctor_name"]} on '
                   f'{appt["appointment_date"]} at {appt["appointment_time"]} has been confirmed.')
            title = '✅ Appointment Confirmed'
            
            # Send acceptance email to patient
            mode = appt['mode'] if 'mode' in appt.keys() and appt['mode'] else 'in-person'
            print(f"[API] Sending acceptance email to {appt['patient_email']}")
            send_appointment_accepted_email(
                patient_email=appt['patient_email'],
                patient_name=appt['patient_name'],
                doctor_name=appt['doctor_name'],
                date=appt['appointment_date'],
                time=appt['appointment_time'],
                mode=mode,
                appointment_id=appointment_id
            )
        else:
            msg = (f'Your appointment with Dr. {appt["doctor_name"]} on '
                   f'{appt["appointment_date"]} at {appt["appointment_time"]} has been rejected.')
            title = '❌ Appointment Rejected'

        conn.execute('''INSERT INTO notifications (user_id, title, message, type)
                        VALUES (?, ?, ?, ?)''',
                     (appt['patient_user_id'], title, msg, 'appointment'))
        conn.commit()
        conn.close()

        print(f"[API] ✅ Appointment #{appointment_id} {action}ed successfully")
        return jsonify({'success': True, 'status': new_status})

    except Exception as e:
        print(f"[API] ❌ Error: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)})


# ==================== VIDEO CALL (Jitsi Meet) ====================

@app.route('/video-call/<int:appointment_id>')
@login_required
def video_call(appointment_id):
    """
    Launch a Jitsi Meet video call room for a confirmed appointment.
    Room name is derived from appointment ID — same URL for doctor and patient.
    """
    conn = get_db()
    appt = conn.execute('''
        SELECT a.*,
               u1.name as patient_name,
               u2.name as doctor_name
        FROM appointments a
        JOIN patients p  ON a.patient_id = p.id
        JOIN users u1    ON p.user_id = u1.id
        JOIN doctors d   ON a.doctor_id = d.id
        JOIN users u2    ON d.user_id = u2.id
        WHERE a.id = ?
    ''', (appointment_id,)).fetchone()
    conn.close()

    if not appt:
        flash('Appointment not found', 'danger')
        if session.get('role') == 'doctor':
            return redirect(url_for('doctor_appointments'))
        return redirect(url_for('patient_appointments'))

    # Unique room name — same for both doctor and patient
    room_name = f"MedicalSystem-Appt-{appointment_id}"
    display_name = session.get('name', 'User')

    return render_template('video_call.html',
                           appointment=appt,
                           room_name=room_name,
                           display_name=display_name)


# ==================== PRESCRIPTION & LAB TESTS (After Video Call) ====================

@app.route('/appointment/cancel/<int:appointment_id>')
@login_required
def cancel_appointment(appointment_id):
    conn = get_db()
    
    # Get appointment details for notification
    appointment = conn.execute('''SELECT a.*, 
                                 u1.name as patient_name, u1.id as patient_user_id,
                                 u2.name as doctor_name, u2.id as doctor_user_id
                                 FROM appointments a
                                 JOIN patients p ON a.patient_id = p.id
                                 JOIN users u1 ON p.user_id = u1.id
                                 JOIN doctors d ON a.doctor_id = d.id
                                 JOIN users u2 ON d.user_id = u2.id
                                 WHERE a.id = ?''', (appointment_id,)).fetchone()
    
    if appointment:
        conn.execute('UPDATE appointments SET status = "cancelled" WHERE id = ?', (appointment_id,))
        
        # Send notification to both parties
        if session['role'] == 'patient':
            # Notify doctor
            conn.execute('''INSERT INTO notifications (user_id, title, message, type)
                           VALUES (?, ?, ?, ?)''',
                        (appointment['doctor_user_id'],
                         'Appointment Cancelled',
                         f'Patient {appointment["patient_name"]} cancelled appointment on {appointment["appointment_date"]} at {appointment["appointment_time"]}',
                         'appointment'))
        else:
            # Notify patient
            conn.execute('''INSERT INTO notifications (user_id, title, message, type)
                           VALUES (?, ?, ?, ?)''',
                        (appointment['patient_user_id'],
                         'Appointment Cancelled',
                         f'Dr. {appointment["doctor_name"]} cancelled your appointment on {appointment["appointment_date"]} at {appointment["appointment_time"]}',
                         'appointment'))
        
        conn.commit()
        flash('Appointment cancelled successfully!', 'info')
    else:
        flash('Appointment not found', 'danger')
    
    conn.close()
    return redirect(request.referrer or url_for('patient_dashboard'))


# End Video Consultation (Mark as Completed)
@app.route('/api/end-consultation/<int:appointment_id>', methods=['POST'])
@login_required
def end_consultation(appointment_id):
    """Mark appointment as completed when video call ends and create consultation record"""
    conn = get_db()
    
    try:
        # Get appointment details
        appointment = conn.execute('''SELECT a.*, 
                                     u1.name as patient_name, u1.id as patient_user_id,
                                     u2.name as doctor_name, u2.id as doctor_user_id
                                     FROM appointments a
                                     JOIN patients p ON a.patient_id = p.id
                                     JOIN users u1 ON p.user_id = u1.id
                                     JOIN doctors d ON a.doctor_id = d.id
                                     JOIN users u2 ON d.user_id = u2.id
                                     WHERE a.id = ?''', (appointment_id,)).fetchone()
        
        if not appointment:
            conn.close()
            return jsonify({'success': False, 'error': 'Appointment not found'}), 404
        
        # Check if appointment is already completed to prevent duplicates
        if appointment['status'] == 'completed':
            # Check if consultation record already exists
            existing_consultation = conn.execute('''
                SELECT id FROM consultations 
                WHERE patient_id = ? AND doctor_id = ? 
                AND predicted_disease = 'Video Consultation'
                AND consultation_date > datetime('now', '-1 hour')
            ''', (appointment['patient_id'], appointment['doctor_id'])).fetchone()
            
            if existing_consultation:
                conn.close()
                return jsonify({'success': True, 'message': 'Consultation already completed', 'duplicate': True})
        
        # Update appointment status to completed
        conn.execute('''UPDATE appointments 
                       SET status = "completed" 
                       WHERE id = ?''', (appointment_id,))
        
        # Check again if consultation record already exists (race condition prevention)
        existing_consultation = conn.execute('''
            SELECT id FROM consultations 
            WHERE patient_id = ? AND doctor_id = ? 
            AND predicted_disease = 'Video Consultation'
            AND consultation_date > datetime('now', '-1 hour')
        ''', (appointment['patient_id'], appointment['doctor_id'])).fetchone()
        
        if not existing_consultation:
            # Create consultation record for completed video conference
            conn.execute('''INSERT INTO consultations 
                           (patient_id, doctor_id, symptoms, diagnosis, prescription, status, consultation_date, predicted_disease)
                           VALUES (?, ?, ?, ?, ?, ?, ?, ?)''',
                        (appointment['patient_id'],
                         appointment['doctor_id'],
                         appointment['reason'] or 'Video consultation completed',
                         appointment['notes'] or 'Consultation completed via video call',
                         appointment['prescription'] or '',
                         'completed',
                         datetime.now().isoformat(),
                         'Video Consultation'))
        
        # Notify patient that consultation is completed (avoid duplicates)
        if session.get('role') == 'doctor':
            # Check if notification already exists
            existing_notification = conn.execute('''
                SELECT id FROM notifications 
                WHERE user_id = ? AND title = ? AND message LIKE ?
                AND created_at > datetime('now', '-1 hour')
            ''', (appointment['patient_user_id'], 
                  'Video Consultation Completed',
                  f'%{appointment["doctor_name"]}%')).fetchone()
            
            if not existing_notification:
                conn.execute('''INSERT INTO notifications (user_id, title, message, type)
                               VALUES (?, ?, ?, ?)''',
                            (appointment['patient_user_id'],
                             'Video Consultation Completed',
                             f'Your video consultation with Dr. {appointment["doctor_name"]} has been completed. Check your consultation history for prescription and feedback options.',
                             'consultation'))
        
        conn.commit()
        conn.close()
        return jsonify({'success': True, 'message': 'Video consultation completed and added to consultation history'})
    
    except Exception as e:
        conn.close()
        return jsonify({'success': False, 'error': str(e)}), 500


# Delete Appointment (Permanently)
@app.route('/appointment/delete/<int:appointment_id>')
@login_required
def delete_appointment(appointment_id):
    conn = get_db()
    
    # Get appointment details to verify ownership
    appointment = conn.execute('''SELECT a.*, p.user_id as patient_user_id, d.user_id as doctor_user_id
                                 FROM appointments a
                                 JOIN patients p ON a.patient_id = p.id
                                 JOIN doctors d ON a.doctor_id = d.id
                                 WHERE a.id = ?''', (appointment_id,)).fetchone()
    
    if not appointment:
        flash('Appointment not found', 'danger')
        conn.close()
        return redirect(request.referrer or url_for('patient_dashboard'))
    
    # Check if user has permission to delete (only patient or doctor involved, or admin)
    if session['role'] == 'patient' and appointment['patient_user_id'] != session['user_id']:
        flash('You do not have permission to delete this appointment', 'danger')
        conn.close()
        return redirect(request.referrer or url_for('patient_dashboard'))
    
    if session['role'] == 'doctor' and appointment['doctor_user_id'] != session['user_id']:
        flash('You do not have permission to delete this appointment', 'danger')
        conn.close()
        return redirect(request.referrer or url_for('doctor_dashboard'))
    
    # Only allow deletion of cancelled or completed appointments
    if appointment['status'] not in ['cancelled', 'completed']:
        flash('Only cancelled or completed appointments can be deleted. Please cancel the appointment first.', 'warning')
        conn.close()
        return redirect(request.referrer or url_for('patient_dashboard'))
    
    # Delete the appointment
    conn.execute('DELETE FROM appointments WHERE id = ?', (appointment_id,))
    conn.commit()
    conn.close()
    
    flash('Appointment deleted successfully!', 'success')
    
    # Redirect based on role
    if session['role'] == 'patient':
        return redirect(url_for('patient_appointments'))
    elif session['role'] == 'doctor':
        return redirect(url_for('doctor_appointments'))
    else:
        return redirect(url_for('admin_dashboard'))

# Reschedule Appointment
@app.route('/appointment/reschedule/<int:appointment_id>', methods=['GET', 'POST'])
@login_required
def reschedule_appointment(appointment_id):
    conn = get_db()
    
    if request.method == 'POST':
        new_date = request.form['appointment_date']
        new_time = request.form['appointment_time']
        
        # Get appointment details for notification
        appointment = conn.execute('''SELECT a.*, 
                                     u1.name as patient_name, u1.id as patient_user_id,
                                     u2.name as doctor_name, u2.id as doctor_user_id
                                     FROM appointments a
                                     JOIN patients p ON a.patient_id = p.id
                                     JOIN users u1 ON p.user_id = u1.id
                                     JOIN doctors d ON a.doctor_id = d.id
                                     JOIN users u2 ON d.user_id = u2.id
                                     WHERE a.id = ?''', (appointment_id,)).fetchone()
        
        # Update appointment
        conn.execute('''UPDATE appointments 
                       SET appointment_date = ?, appointment_time = ?, status = 'rescheduled'
                       WHERE id = ?''',
                    (new_date, new_time, appointment_id))
        
        # Send notifications
        if session['role'] == 'patient':
            conn.execute('''INSERT INTO notifications (user_id, title, message, type)
                           VALUES (?, ?, ?, ?)''',
                        (appointment['doctor_user_id'],
                         'Appointment Rescheduled',
                         f'Patient {appointment["patient_name"]} rescheduled appointment to {new_date} at {new_time}',
                         'appointment'))
        else:
            conn.execute('''INSERT INTO notifications (user_id, title, message, type)
                           VALUES (?, ?, ?, ?)''',
                        (appointment['patient_user_id'],
                         'Appointment Rescheduled',
                         f'Dr. {appointment["doctor_name"]} rescheduled your appointment to {new_date} at {new_time}',
                         'appointment'))
        
        conn.commit()
        conn.close()
        
        flash('Appointment rescheduled successfully!', 'success')
        return redirect(request.referrer or url_for('patient_dashboard'))
    
    # GET request - show reschedule form
    appointment = conn.execute('''SELECT a.*, d.specialization, u.name as doctor_name
                                 FROM appointments a
                                 JOIN doctors d ON a.doctor_id = d.id
                                 JOIN users u ON d.user_id = u.id
                                 WHERE a.id = ?''', (appointment_id,)).fetchone()
    
    conn.close()
    
    from datetime import date
    today = date.today().strftime('%Y-%m-%d')
    
    return render_template('reschedule_appointment.html', appointment=appointment, today=today)

# Add Appointment Notes
@app.route('/appointment/add-note/<int:appointment_id>', methods=['POST'])
@login_required
def add_appointment_note(appointment_id):
    note = request.form.get('note', '')
    
    conn = get_db()
    conn.execute('UPDATE appointments SET notes = ? WHERE id = ?', (note, appointment_id))
    conn.commit()
    conn.close()
    
    flash('Note added successfully!', 'success')
    return redirect(request.referrer or url_for('patient_dashboard'))

# View Prescription (Patient Side)
@app.route('/appointment/view-prescription/<int:appointment_id>')
@login_required
@role_required('patient')
def view_prescription(appointment_id):
    """Patient view of prescription with nearby services"""
    conn = get_db()
    
    patient = conn.execute('SELECT * FROM patients WHERE user_id = ?', (session['user_id'],)).fetchone()
    
    appointment = conn.execute('''
        SELECT a.*, d.specialization, u.name as doctor_name
        FROM appointments a
        JOIN doctors d ON a.doctor_id = d.id
        JOIN users u ON d.user_id = u.id
        WHERE a.id = ? AND a.patient_id = ?
    ''', (appointment_id, patient['id'])).fetchone()
    
    conn.close()
    
    if not appointment:
        flash('Appointment not found', 'danger')
        return redirect(url_for('patient_appointments'))
    
    if not (appointment['prescription'] or appointment['lab_tests']):
        flash('No prescription available for this appointment yet', 'info')
        return redirect(url_for('patient_appointments'))
    
    return render_template('view_prescription_modern.html', appointment=appointment)

# Add/Update Prescription for Appointment
@app.route('/appointment/add-prescription/<int:appointment_id>', methods=['GET', 'POST'])
@login_required
@role_required('doctor')
def add_prescription(appointment_id):
    """Add or update prescription and lab tests for an appointment"""
    conn = get_db()
    
    if request.method == 'POST':
        prescription = request.form.get('prescription', '')
        lab_tests = request.form.get('lab_tests', '')
        
        # Update appointment with prescription and lab tests
        conn.execute('''UPDATE appointments 
                       SET prescription = ?, lab_tests = ?, status = 'completed'
                       WHERE id = ?''',
                    (prescription, lab_tests, appointment_id))
        
        # Get appointment details for notification
        appointment = conn.execute('''
            SELECT a.*, p.user_id as patient_user_id, u.name as patient_name
            FROM appointments a
            JOIN patients p ON a.patient_id = p.id
            JOIN users u ON p.user_id = u.id
            WHERE a.id = ?
        ''', (appointment_id,)).fetchone()
        
        if appointment and (prescription or lab_tests):
            # Determine service type
            service_type = 'all'
            if prescription and not lab_tests:
                service_type = 'pharmacy'
            elif lab_tests and not prescription:
                service_type = 'diagnostic'
            
            # Create notification for patient
            if service_type == 'pharmacy':
                notification_title = '💊 Prescription Ready'
                notification_message = f'Dr. {session["name"]} has prescribed medicines. Click to find nearby pharmacies.'
            elif service_type == 'diagnostic':
                notification_title = '🔬 Lab Tests Recommended'
                notification_message = f'Dr. {session["name"]} has recommended lab tests. Click to find nearby diagnostic centers.'
            else:
                notification_title = '🏥 Prescription & Tests Ready'
                notification_message = f'Dr. {session["name"]} has prescribed medicines and tests. Click to find nearby services.'
            
            conn.execute('''INSERT INTO notifications (user_id, title, message, type)
                           VALUES (?, ?, ?, ?)''',
                        (appointment['patient_user_id'], notification_title, notification_message, 'prescription'))
        
        conn.commit()
        conn.close()
        
        flash('Prescription added successfully!', 'success')
        return redirect(url_for('doctor_appointments'))
    
    # GET request - show form
    appointment = conn.execute('''
        SELECT a.*, u.name as patient_name, u.email as patient_email
        FROM appointments a
        JOIN patients p ON a.patient_id = p.id
        JOIN users u ON p.user_id = u.id
        WHERE a.id = ?
    ''', (appointment_id,)).fetchone()
    
    conn.close()
    
    if not appointment:
        flash('Appointment not found', 'danger')
        return redirect(url_for('doctor_appointments'))
    
    # Render the prescription form
    return render_template('add_prescription_modern.html', appointment=appointment)

# ==================== FEEDBACK & RATINGS ====================

@app.route('/patient/feedback/<int:consultation_id>', methods=['GET', 'POST'])
@login_required
@role_required('patient')
def submit_feedback(consultation_id):
    if request.method == 'POST':
        rating = request.form['rating']
        comment = request.form['comment']
        
        conn = get_db()
        patient = conn.execute('SELECT * FROM patients WHERE user_id = ?', (session['user_id'],)).fetchone()
        consultation = conn.execute('SELECT doctor_id FROM consultations WHERE id = ?', (consultation_id,)).fetchone()
        
        conn.execute('''INSERT INTO feedback (patient_id, doctor_id, consultation_id, rating, comment)
                       VALUES (?, ?, ?, ?, ?)''',
                    (patient['id'], consultation['doctor_id'], consultation_id, rating, comment))
        conn.commit()
        conn.close()
        
        flash('Thank you for your feedback!', 'success')
        return redirect(url_for('consultation_history'))
    
    conn = get_db()
    consultation = conn.execute('''SELECT c.*, u.name as doctor_name
                                  FROM consultations c
                                  JOIN doctors d ON c.doctor_id = d.id
                                  JOIN users u ON d.user_id = u.id
                                  WHERE c.id = ?''', (consultation_id,)).fetchone()
    conn.close()
    
    return render_template('submit_feedback.html', consultation=consultation)

@app.route('/doctor/feedback')
@login_required
@role_required('doctor')
def doctor_feedback():
    conn = get_db()
    doctor = conn.execute('SELECT * FROM doctors WHERE user_id = ?', (session['user_id'],)).fetchone()
    
    feedbacks = conn.execute('''SELECT f.*, u.name as patient_name
                               FROM feedback f
                               JOIN patients p ON f.patient_id = p.id
                               JOIN users u ON p.user_id = u.id
                               WHERE f.doctor_id = ?
                               ORDER BY f.created_at DESC''',
                            (doctor['id'],)).fetchall()
    
    # Calculate average rating
    avg_rating = conn.execute('''SELECT AVG(rating) as avg_rating 
                                FROM feedback 
                                WHERE doctor_id = ?''', (doctor['id'],)).fetchone()
    
    conn.close()
    
    return render_template('doctor_feedback.html', feedbacks=feedbacks, avg_rating=avg_rating['avg_rating'] or 0)

# ==================== MEDICAL RECORDS ====================

@app.route('/patient/medical-records')
@login_required
@role_required('patient')
def patient_medical_records():
    conn = get_db()
    patient = conn.execute('SELECT * FROM patients WHERE user_id = ?', (session['user_id'],)).fetchone()
    
    records = conn.execute('''SELECT m.*, u.name as uploaded_by_name
                             FROM medical_records m
                             LEFT JOIN users u ON m.uploaded_by = u.id
                             WHERE m.patient_id = ?
                             ORDER BY m.created_at DESC''',
                          (patient['id'],)).fetchall()
    conn.close()
    
    return render_template('patient_medical_records.html', records=records)

@app.route('/patient/add-medical-record', methods=['POST'])
@login_required
@role_required('patient')
def add_medical_record():
    record_type = request.form['record_type']
    record_title = request.form['record_title']
    record_data = request.form['record_data']
    
    # Handle file upload
    file_path = None
    if 'record_file' in request.files:
        file = request.files['record_file']
        if file and file.filename != '':
            # Secure the filename
            from werkzeug.utils import secure_filename
            import os
            
            filename = secure_filename(file.filename)
            # Add timestamp to filename to make it unique
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"{timestamp}_{filename}"
            
            # Save file
            upload_folder = 'static/uploads'
            if not os.path.exists(upload_folder):
                os.makedirs(upload_folder)
            
            file_path = os.path.join(upload_folder, filename)
            file.save(file_path)
            
            # Store relative path
            file_path = f"uploads/{filename}"
    
    conn = get_db()
    patient = conn.execute('SELECT * FROM patients WHERE user_id = ?', (session['user_id'],)).fetchone()
    
    # Store file path in record_data if file was uploaded
    if file_path:
        record_data = f"{record_data}\n\n[File: {file_path}]"
    
    conn.execute('''INSERT INTO medical_records (patient_id, record_type, record_title, record_data, uploaded_by)
                   VALUES (?, ?, ?, ?, ?)''',
                (patient['id'], record_type, record_title, record_data, session['user_id']))
    conn.commit()
    conn.close()
    
    flash('Medical record added successfully!', 'success')
    return redirect(url_for('patient_medical_records'))

# ==================== NOTIFICATIONS ====================

@app.route('/notifications')
@login_required
def notifications():
    conn = get_db()
    notifications = conn.execute('''SELECT * FROM notifications 
                                   WHERE user_id = ? 
                                   ORDER BY created_at DESC 
                                   LIMIT 50''',
                                (session['user_id'],)).fetchall()
    
    # Mark as read
    conn.execute('UPDATE notifications SET is_read = 1 WHERE user_id = ?', (session['user_id'],))
    conn.commit()
    conn.close()
    
    return render_template('notifications.html', notifications=notifications)

@app.route('/notifications/count')
@login_required
def notification_count():
    conn = get_db()
    count = conn.execute('''SELECT COUNT(*) as count FROM notifications 
                           WHERE user_id = ? AND is_read = 0''',
                        (session['user_id'],)).fetchone()
    conn.close()
    
    return jsonify({'count': count['count']})

# ==================== DOCTOR AVAILABILITY ====================

@app.route('/doctor/availability', methods=['GET', 'POST'])
@login_required
@role_required('doctor')
def doctor_availability():
    conn = get_db()
    doctor = conn.execute('SELECT * FROM doctors WHERE user_id = ?', (session['user_id'],)).fetchone()
    
    if request.method == 'POST':
        day_of_week = request.form['day_of_week']
        start_time = request.form['start_time']
        end_time = request.form['end_time']
        
        conn.execute('''INSERT INTO doctor_availability (doctor_id, day_of_week, start_time, end_time)
                       VALUES (?, ?, ?, ?)''',
                    (doctor['id'], day_of_week, start_time, end_time))
        conn.commit()
        flash('Availability added successfully!', 'success')
    
    availability = conn.execute('''SELECT * FROM doctor_availability 
                                  WHERE doctor_id = ?
                                  ORDER BY 
                                    CASE day_of_week
                                        WHEN 'Monday' THEN 1
                                        WHEN 'Tuesday' THEN 2
                                        WHEN 'Wednesday' THEN 3
                                        WHEN 'Thursday' THEN 4
                                        WHEN 'Friday' THEN 5
                                        WHEN 'Saturday' THEN 6
                                        WHEN 'Sunday' THEN 7
                                    END''',
                               (doctor['id'],)).fetchall()
    conn.close()
    
    return render_template('doctor_availability.html', availability=availability)

@app.route('/doctor/availability/delete/<int:availability_id>')
@login_required
@role_required('doctor')
def delete_availability(availability_id):
    conn = get_db()
    conn.execute('DELETE FROM doctor_availability WHERE id = ?', (availability_id,))
    conn.commit()
    conn.close()
    
    flash('Availability deleted successfully!', 'info')
    return redirect(url_for('doctor_availability'))

# ==================== PRESCRIPTION PDF ====================

@app.route('/prescription/download/<int:consultation_id>')
@login_required
def download_prescription(consultation_id):
    try:
        from reportlab.lib.pagesizes import letter
        from reportlab.pdfgen import canvas
        from io import BytesIO
        from flask import send_file
        
        conn = get_db()
        consultation = conn.execute('''SELECT c.*, 
                                      u1.name as patient_name, u1.email as patient_email,
                                      u2.name as doctor_name, d.specialization, d.qualification
                                      FROM consultations c
                                      JOIN patients p ON c.patient_id = p.id
                                      JOIN users u1 ON p.user_id = u1.id
                                      JOIN doctors d ON c.doctor_id = d.id
                                      JOIN users u2 ON d.user_id = u2.id
                                      WHERE c.id = ?''', (consultation_id,)).fetchone()
        
        # If consultation not found, check appointments table
        if not consultation:
            consultation = conn.execute('''SELECT a.*, 
                                          u1.name as patient_name, u1.email as patient_email,
                                          u2.name as doctor_name, d.specialization, d.qualification,
                                          a.created_at as consultation_date
                                          FROM appointments a
                                          JOIN patients p ON a.patient_id = p.id
                                          JOIN users u1 ON p.user_id = u1.id
                                          JOIN doctors d ON a.doctor_id = d.id
                                          JOIN users u2 ON d.user_id = u2.id
                                          WHERE a.id = ?''', (consultation_id,)).fetchone()
        
        conn.close()
        
        if not consultation:
            flash('Prescription not found', 'error')
            return redirect(request.referrer or url_for('patient_dashboard'))
        
        # Create PDF
        buffer = BytesIO()
        p = canvas.Canvas(buffer, pagesize=letter)
        
        # Add content
        p.setFont("Helvetica-Bold", 20)
        p.drawString(100, 750, "Medical Prescription")
        
        p.setFont("Helvetica", 12)
        consultation_date = consultation['consultation_date'] or consultation.get('created_at', '')
        if consultation_date:
            date_str = consultation_date[:10] if len(consultation_date) >= 10 else consultation_date
            p.drawString(100, 700, f"Date: {date_str}")
        else:
            p.drawString(100, 700, "Date: N/A")
            
        p.drawString(100, 680, f"Doctor: Dr. {consultation['doctor_name']}")
        p.drawString(100, 660, f"Specialization: {consultation['specialization']}")
        p.drawString(100, 640, f"Patient: {consultation['patient_name']}")
        
        p.drawString(100, 600, "Diagnosis:")
        diagnosis = consultation.get('diagnosis') or consultation.get('notes') or 'N/A'
        p.drawString(120, 580, diagnosis)
        
        p.drawString(100, 540, "Prescription:")
        prescription = consultation.get('prescription') or 'N/A'
        p.drawString(120, 520, prescription)
        
        p.showPage()
        p.save()
        
        buffer.seek(0)
        return send_file(buffer, as_attachment=True, download_name=f'prescription_{consultation_id}.pdf', mimetype='application/pdf')
    except ImportError:
        flash('PDF generation requires reportlab. Install it with: pip install reportlab', 'warning')
        return redirect(request.referrer or url_for('patient_dashboard'))
    except Exception as e:
        flash(f'Error generating prescription: {str(e)}', 'error')
        return redirect(request.referrer or url_for('patient_dashboard'))

# ==================== EMERGENCY CONTACT ====================

@app.route('/emergency')
def emergency():
    return render_template('emergency.html')

# ==================== PUSH NOTIFICATION API ROUTES ====================

# SSE stream — doctor's browser connects once and receives live events
@app.route('/api/doctor/stream')
@login_required
@role_required('doctor')
def doctor_notification_stream():
    """
    Server-Sent Events endpoint.
    Doctor dashboard connects here; when a patient books an appointment
    push_notification_to_doctor() puts data in the queue and this
    streams it to the browser instantly.
    """
    user_id = session['user_id']
    q = get_doctor_queue(user_id)

    def event_stream():
        # Send a heartbeat every 20s to keep connection alive
        while True:
            try:
                data = q.get(timeout=20)
                yield f"data: {json.dumps(data)}\n\n"
            except queue.Empty:
                yield "data: {\"type\":\"heartbeat\"}\n\n"

    return Response(
        event_stream(),
        mimetype='text/event-stream',
        headers={
            'Cache-Control':   'no-cache',
            'X-Accel-Buffering': 'no',
            'Connection':      'keep-alive',
        }
    )

# API: New appointments for doctor (polled every 8s by dashboard)
@app.route('/api/doctor/new-appointments')
@login_required
@role_required('doctor')
def api_doctor_new_appointments():
    """
    Returns appointments created in the last 30 seconds for this doctor.
    The doctor dashboard polls this and shows a live alert popup.
    """
    conn = get_db()
    doctor = conn.execute('SELECT id FROM doctors WHERE user_id = ?', (session['user_id'],)).fetchone()
    if not doctor:
        conn.close()
        return jsonify({'appointments': []})

    new_appts = conn.execute('''
        SELECT a.*, u.name as patient_name, u.email as patient_email
        FROM appointments a
        JOIN patients p ON a.patient_id = p.id
        JOIN users u ON p.user_id = u.id
        WHERE a.doctor_id = ?
          AND datetime(a.created_at) >= datetime('now', '-30 seconds')
        ORDER BY a.created_at DESC
    ''', (doctor['id'],)).fetchall()
    conn.close()

    return jsonify({'appointments': [{
        'id':               r['id'],
        'patient_name':     r['patient_name'],
        'patient_email':    r['patient_email'],
        'appointment_date': r['appointment_date'],
        'appointment_time': r['appointment_time'],
        'reason':           r['reason'] or '',
        'status':           r['status'],
        'is_emergency':     r['status'] == 'emergency',
    } for r in new_appts]})

# API: Get latest unread notifications
@app.route('/api/notifications/latest')
@login_required
def api_notifications_latest():
    conn = get_db()
    
    # Get unread notifications from last 5 minutes
    notifications = conn.execute('''
        SELECT * FROM notifications 
        WHERE user_id = ? 
        AND is_read = 0 
        AND datetime(created_at) > datetime('now', '-5 minutes')
        ORDER BY created_at DESC
    ''', (session['user_id'],)).fetchall()
    
    conn.close()
    
    return jsonify({
        'notifications': [{
            'id': n['id'],
            'title': n['title'],
            'message': n['message'],
            'type': n['type'],
            'created_at': n['created_at']
        } for n in notifications]
    })

# API: Mark notification as read
@app.route('/api/notifications/<int:notification_id>/read', methods=['POST'])
@login_required
def api_mark_notification_read(notification_id):
    conn = get_db()
    conn.execute('UPDATE notifications SET is_read = 1 WHERE id = ? AND user_id = ?',
                (notification_id, session['user_id']))
    conn.commit()
    conn.close()
    
    return jsonify({'success': True})

# API: Mark all notifications as read
@app.route('/api/notifications/read-all', methods=['POST'])
@login_required
def api_mark_all_notifications_read():
    conn = get_db()
    conn.execute('UPDATE notifications SET is_read = 1 WHERE user_id = ?',
                (session['user_id'],))
    conn.commit()
    conn.close()
    
    return jsonify({'success': True})

# Test notification route
@app.route('/test-notification')
@login_required
def test_notification():
    conn = get_db()
    conn.execute('''INSERT INTO notifications (user_id, title, message, type)
                   VALUES (?, ?, ?, ?)''',
                (session['user_id'],
                 '🔔 Test Notification',
                 'This is a test push notification! If you see this, notifications are working perfectly.',
                 'info'))
    conn.commit()
    conn.close()
    
    flash('Test notification created! Wait 30 seconds or refresh to see the browser notification.', 'success')
    return redirect(url_for('notifications'))

# ==================== SOCKET.IO EVENT HANDLERS ====================

if SOCKETIO_AVAILABLE and socketio:
    @socketio.on('join_doctor_room')
    def on_join_doctor_room(data):
        """Doctor joins their personal notification room on dashboard load."""
        room = data.get('room')
        if room:
            join_room(room)

    @socketio.on('leave_doctor_room')
    def on_leave_doctor_room(data):
        room = data.get('room')
        if room:
            leave_room(room)
    
    # ==================== VIDEO CALL SYNCHRONIZATION ====================
    
    @socketio.on('join_video_call')
    def on_join_video_call(data):
        """User joins video call room for real-time synchronization"""
        appointment_id = data.get('appointment_id')
        user_role = data.get('role')  # 'doctor' or 'patient'
        
        if appointment_id:
            room = f"video_call_{appointment_id}"
            join_room(room)
            
            # Notify other party that someone joined
            emit('user_joined_call', {
                'role': user_role,
                'message': f'{user_role.title()} joined the video call'
            }, room=room, include_self=False)
    
    @socketio.on('end_video_call')
    def on_end_video_call(data):
        """Handle video call ending - notify both parties"""
        appointment_id = data.get('appointment_id')
        user_role = data.get('role')
        
        if appointment_id:
            room = f"video_call_{appointment_id}"
            
            # Notify all users in the room that call is ending
            emit('call_ended_by_peer', {
                'ended_by': user_role,
                'message': f'Call ended by {user_role}',
                'appointment_id': appointment_id
            }, room=room)
            
            # Leave the room
            leave_room(room)


# ==================== NEARBY SERVICES - SMART SUGGESTIONS ====================

@app.route('/nearby-services')
@login_required
def nearby_services_general():
    """
    General nearby services page - accessible from prescriptions
    """
    service_type = request.args.get('type', 'all')
    appointment_id = request.args.get('appointment_id', '')
    
    # Create a simple consultation-like object for the template
    consultation_data = {
        'id': 0,
        'doctor_name': 'Your Doctor',
        'specialization': '',
        'consultation_date': '',
        'diagnosis': '',
        'prescription': ''
    }
    
    return render_template('nearby_services.html', 
                         service_type=service_type,
                         appointment_id=appointment_id,
                         consultation=consultation_data)

@app.route('/patient/nearby-services/<int:consultation_id>')
@login_required
@role_required('patient')
def nearby_services_page(consultation_id):
    """
    Display nearby medical stores and diagnostic labs based on prescription
    """
    conn = get_db()
    
    # First try to get consultation details
    consultation = conn.execute('''
        SELECT c.*, u.name as doctor_name, d.specialization,
               u2.name as patient_name, u2.address as patient_address
        FROM consultations c
        JOIN doctors d ON c.doctor_id = d.id
        JOIN users u ON d.user_id = u.id
        JOIN patients p ON c.patient_id = p.id
        JOIN users u2 ON p.user_id = u2.id
        WHERE c.id = ?
    ''', (consultation_id,)).fetchone()
    
    # If no consultation found, try appointments table (consultation_id might be appointment_id)
    if not consultation:
        appointment = conn.execute('''
            SELECT a.id, a.prescription, a.lab_tests, a.appointment_date as consultation_date,
                   a.reason as symptoms, a.notes as diagnosis, a.status,
                   u.name as doctor_name, d.specialization,
                   u2.name as patient_name, u2.address as patient_address
            FROM appointments a
            JOIN doctors d ON a.doctor_id = d.id
            JOIN users u ON d.user_id = u.id
            JOIN patients p ON a.patient_id = p.id
            JOIN users u2 ON p.user_id = u2.id
            WHERE a.id = ?
        ''', (consultation_id,)).fetchone()
        
        # Convert appointment to consultation format for template compatibility
        if appointment:
            consultation = {
                'id': appointment['id'],
                'prescription': appointment['prescription'],
                'lab_tests': appointment['lab_tests'],
                'consultation_date': appointment['consultation_date'],
                'symptoms': appointment['symptoms'],
                'diagnosis': appointment['diagnosis'],
                'status': appointment['status'],
                'doctor_name': appointment['doctor_name'],
                'specialization': appointment['specialization'],
                'patient_name': appointment['patient_name'],
                'patient_address': appointment['patient_address']
            }
    
    conn.close()
    
    if not consultation:
        flash('Consultation/Appointment not found', 'danger')
        return redirect(url_for('consultation_history'))
    
    # Analyze prescription to determine service type
    prescription = consultation['prescription'] or ''
    service_type = analyze_prescription_type(prescription)
    
    return render_template('nearby_services.html', 
                         consultation=consultation,
                         service_type=service_type)

# Direct appointment prescription view
@app.route('/patient/prescription/<int:appointment_id>')
@login_required
@role_required('patient')
def view_appointment_prescription(appointment_id):
    """
    View prescription from appointment (alternative to consultation-based view)
    """
    conn = get_db()
    
    # Get appointment with prescription
    appointment = conn.execute('''
        SELECT a.*, u.name as doctor_name, d.specialization,
               u2.name as patient_name, u2.address as patient_address
        FROM appointments a
        JOIN doctors d ON a.doctor_id = d.id
        JOIN users u ON d.user_id = u.id
        JOIN patients p ON a.patient_id = p.id
        JOIN users u2 ON p.user_id = u2.id
        WHERE a.id = ? AND p.user_id = ?
    ''', (appointment_id, session['user_id'])).fetchone()
    
    conn.close()
    
    if not appointment:
        flash('Appointment not found', 'danger')
        return redirect(url_for('patient_appointments'))
    
    if not (appointment['prescription'] or appointment['lab_tests']):
        flash('No prescription available for this appointment yet', 'info')
        return redirect(url_for('patient_appointments'))
    
    # Use the appointment data as consultation data for the template
    consultation_data = {
        'id': appointment['id'],
        'prescription': appointment['prescription'],
        'lab_tests': appointment['lab_tests'],
        'consultation_date': appointment['appointment_date'],
        'symptoms': appointment['reason'],
        'diagnosis': appointment['notes'],
        'status': appointment['status'],
        'doctor_name': appointment['doctor_name'],
        'specialization': appointment['specialization'],
        'patient_name': appointment['patient_name'],
        'patient_address': appointment['patient_address']
    }
    
    # Analyze prescription to determine service type
    prescription = appointment['prescription'] or ''
    service_type = analyze_prescription_type(prescription)
    
    return render_template('nearby_services.html', 
                         consultation=consultation_data,
                         service_type=service_type)


@app.route('/api/nearby-services', methods=['GET', 'POST'])
@login_required
def api_nearby_services():
    """
    API endpoint to fetch nearby medical services
    Supports both GET (from consultation_history.html) and POST (from nearby_services.html)
    Returns pharmacies and/or diagnostic centers based on prescription
    """
    try:
        # Handle both GET and POST requests
        if request.method == 'GET':
            # GET request from consultation_history.html
            lat = request.args.get('lat', type=float)
            lng = request.args.get('lng', type=float)
            service_type = request.args.get('type', 'all')
            location = f"{lat},{lng}" if lat and lng else None
            prescription = ''
        else:
            # POST request from nearby_services.html
            data = request.get_json()
            location = data.get('location', '')
            prescription = data.get('prescription', '')
            service_type = data.get('service_type', 'all')
        
        if not location:
            return jsonify({'success': False, 'error': 'Location is required', 'results': []})
        
        # Parse location
        if ',' in location:
            try:
                lat, lng = map(float, location.split(','))
            except:
                return jsonify({'success': False, 'error': 'Invalid coordinates', 'results': []})
        else:
            # Try to geocode address
            coords = geocode_address(location)
            if not coords:
                return jsonify({'success': False, 'error': 'Could not find location', 'results': []})
            lat, lng = coords
        
        # Fetch nearby services based on type
        all_results = []
        
        if service_type == 'pharmacy':
            all_results = find_nearby_pharmacies(lat, lng)
        elif service_type == 'lab' or service_type == 'ct_scan' or service_type == 'diagnostic':
            all_results = find_nearby_diagnostic_centers(lat, lng, prescription)
        elif service_type == 'all':
            all_results = find_nearby_pharmacies(lat, lng) + find_nearby_diagnostic_centers(lat, lng, prescription)
        
        # For GET requests (consultation_history.html), return in 'results' format
        if request.method == 'GET':
            return jsonify({
                'results': all_results,
                'status': 'OK'
            })
        
        # For POST requests (nearby_services.html), return in detailed format
        results = {
            'success': True,
            'location': {'lat': lat, 'lng': lng},
            'pharmacies': [],
            'diagnostic_centers': []
        }
        
        if service_type in ['pharmacy', 'all']:
            results['pharmacies'] = find_nearby_pharmacies(lat, lng)
        
        if service_type in ['diagnostic', 'all']:
            results['diagnostic_centers'] = find_nearby_diagnostic_centers(lat, lng, prescription)
        
        return jsonify(results)
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e), 'results': []})


@app.route('/api/analyze-prescription', methods=['POST'])
@login_required
def api_analyze_prescription():
    """
    Analyze prescription text to determine what services are needed
    """
    try:
        data = request.get_json()
        prescription = data.get('prescription', '')
        
        analysis = analyze_prescription_type(prescription)
        medicines = extract_medicines(prescription)
        tests = extract_diagnostic_tests(prescription)
        
        return jsonify({
            'success': True,
            'service_type': analysis,
            'medicines': medicines,
            'tests': tests,
            'suggestions': generate_suggestions(analysis, medicines, tests)
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})


# ==================== HELPER FUNCTIONS FOR NEARBY SERVICES ====================

def analyze_prescription_type(prescription_text):
    """
    Analyze prescription to determine if it contains medicines, tests, or both
    Returns: 'pharmacy', 'diagnostic', or 'both'
    """
    if not prescription_text:
        return 'pharmacy'
    
    text_lower = prescription_text.lower()
    
    # Keywords for diagnostic tests
    diagnostic_keywords = [
        'ct scan', 'ct-scan', 'mri', 'x-ray', 'xray', 'ultrasound', 'sonography',
        'blood test', 'blood work', 'lab test', 'pathology', 'ecg', 'ekg',
        'endoscopy', 'colonoscopy', 'mammography', 'pet scan', 'bone scan',
        'urine test', 'stool test', 'biopsy', 'culture test', 'thyroid test',
        'diabetes test', 'lipid profile', 'liver function', 'kidney function',
        'cbc', 'hemoglobin', 'glucose', 'hba1c', 'tsh', 'vitamin d'
    ]
    
    # Keywords for medicines
    medicine_keywords = [
        'tablet', 'capsule', 'syrup', 'injection', 'mg', 'ml', 'dose',
        'morning', 'evening', 'night', 'before food', 'after food',
        'paracetamol', 'ibuprofen', 'antibiotic', 'medicine', 'medication'
    ]
    
    has_diagnostic = any(keyword in text_lower for keyword in diagnostic_keywords)
    has_medicine = any(keyword in text_lower for keyword in medicine_keywords)
    
    if has_diagnostic and has_medicine:
        return 'both'
    elif has_diagnostic:
        return 'diagnostic'
    else:
        return 'pharmacy'


def extract_medicines(prescription_text):
    """
    Extract medicine names from prescription
    """
    if not prescription_text:
        return []
    
    medicines = []
    lines = prescription_text.split('\n')
    
    # Common medicine patterns
    medicine_indicators = ['tab', 'tablet', 'cap', 'capsule', 'syrup', 'inj', 'injection', 'mg', 'ml']
    
    for line in lines:
        line = line.strip()
        if not line or len(line) < 3:
            continue
        
        line_lower = line.lower()
        
        # Check if line contains medicine indicators
        if any(indicator in line_lower for indicator in medicine_indicators):
            # Extract medicine name (usually the first word or phrase before dosage)
            parts = line.split('-')[0].split('(')[0].strip()
            if parts and len(parts) > 2:
                medicines.append(parts)
    
    return medicines[:10]  # Return max 10 medicines


def extract_diagnostic_tests(prescription_text):
    """
    Extract diagnostic test names from prescription
    """
    if not prescription_text:
        return []
    
    text_lower = prescription_text.lower()
    tests = []
    
    # Common diagnostic tests
    test_patterns = {
        'CT Scan': ['ct scan', 'ct-scan', 'computed tomography'],
        'MRI': ['mri', 'magnetic resonance'],
        'X-Ray': ['x-ray', 'xray', 'radiography'],
        'Ultrasound': ['ultrasound', 'sonography', 'usg'],
        'Blood Test': ['blood test', 'blood work', 'cbc', 'hemoglobin'],
        'ECG': ['ecg', 'ekg', 'electrocardiogram'],
        'Endoscopy': ['endoscopy', 'gastroscopy'],
        'Urine Test': ['urine test', 'urine analysis'],
        'Thyroid Test': ['thyroid', 'tsh', 't3', 't4'],
        'Diabetes Test': ['glucose', 'hba1c', 'diabetes test'],
        'Lipid Profile': ['lipid profile', 'cholesterol'],
        'Liver Function': ['liver function', 'lft', 'sgpt', 'sgot'],
        'Kidney Function': ['kidney function', 'kft', 'creatinine', 'urea']
    }
    
    for test_name, keywords in test_patterns.items():
        if any(keyword in text_lower for keyword in keywords):
            tests.append(test_name)
    
    return tests


def generate_suggestions(service_type, medicines, tests):
    """
    Generate smart suggestions based on prescription analysis
    """
    suggestions = []
    
    if service_type in ['pharmacy', 'both']:
        if medicines:
            suggestions.append({
                'type': 'pharmacy',
                'title': '💊 Buy Prescribed Medicines',
                'message': f'You have {len(medicines)} medicine(s) prescribed. Find nearby medical stores to purchase them.',
                'action': 'Find Medical Stores',
                'icon': '🏪'
            })
        else:
            suggestions.append({
                'type': 'pharmacy',
                'title': '💊 Medicines Prescribed',
                'message': 'Your doctor has prescribed medicines. Find nearby pharmacies to purchase them.',
                'action': 'Find Medical Stores',
                'icon': '🏪'
            })
    
    if service_type in ['diagnostic', 'both']:
        if tests:
            suggestions.append({
                'type': 'diagnostic',
                'title': '🔬 Diagnostic Tests Required',
                'message': f'You need {len(tests)} test(s): {", ".join(tests[:3])}. Find nearby diagnostic centers.',
                'action': 'Find Diagnostic Centers',
                'icon': '🏥'
            })
        else:
            suggestions.append({
                'type': 'diagnostic',
                'title': '🔬 Tests Recommended',
                'message': 'Your doctor has recommended diagnostic tests. Find nearby labs and imaging centers.',
                'action': 'Find Diagnostic Centers',
                'icon': '🏥'
            })
    
    return suggestions


def geocode_address(address):
    """
    Convert address to coordinates using OpenStreetMap Nominatim
    Biased to India for better accuracy with Indian city/pincode searches
    """
    try:
        url = "https://nominatim.openstreetmap.org/search"
        
        # First try with India bias
        params = {
            'q': address,
            'format': 'json',
            'limit': 1,
            'countrycodes': 'in',          # bias to India
            'addressdetails': 1
        }
        headers = {'User-Agent': 'MedicalSystem/1.0 (healthcare app)'}
        
        response = requests.get(url, params=params, headers=headers, timeout=10)
        data = response.json()
        
        if data:
            result = data[0]
            print(f"Geocoded '{address}' → {result.get('display_name', '')} ({result['lat']}, {result['lon']})")
            return (float(result['lat']), float(result['lon']))
        
        # Fallback: try without country restriction
        params.pop('countrycodes', None)
        response = requests.get(url, params=params, headers=headers, timeout=10)
        data = response.json()
        
        if data:
            result = data[0]
            print(f"Geocoded (global) '{address}' → {result.get('display_name', '')} ({result['lat']}, {result['lon']})")
            return (float(result['lat']), float(result['lon']))
        
        return None
    except Exception as e:
        print(f"Geocoding error: {e}")
        return None


def reverse_geocode_city(lat, lng):
    """
    Get city/town name from coordinates using reverse geocoding.
    Returns a human-readable location name.
    """
    try:
        url = "https://nominatim.openstreetmap.org/reverse"
        params = {
            'lat': lat,
            'lon': lng,
            'format': 'json',
            'zoom': 14,
            'addressdetails': 1
        }
        headers = {'User-Agent': 'MedicalSystem/1.0 (healthcare app)'}
        response = requests.get(url, params=params, headers=headers, timeout=5)
        if response.status_code == 200:
            data = response.json()
            addr = data.get('address', {})
            return (addr.get('city') or addr.get('town') or addr.get('village') or
                    addr.get('suburb') or addr.get('county') or "your area")
    except Exception:
        pass
    return "your area"


def find_nearby_pharmacies(lat, lng, radius=5000):
    """
    Find nearby pharmacies - Generates sample data near user's location
    """
    try:
        print(f"Searching for pharmacies near: {lat}, {lng} within {radius}m")
        
        # Use the shared reverse geocoding helper
        location_name = reverse_geocode_city(lat, lng)
        
        import random
        pharmacy_names = ['Apollo Pharmacy', 'MedPlus', 'Wellness Forever', '1mg Store', 'Netmeds', 
                         'PharmEasy', 'Medlife', 'HealthKart', 'Guardian Pharmacy', 'Fortis Pharmacy']
        
        street_names = ['Main Road', 'Gandhi Road', 'MG Road', 'Station Road', 'Market Street', 
                       'Church Street', 'Temple Road', 'Bus Stand Road', 'College Road', 'Hospital Road']
        
        pharmacies = []
        for i in range(5):
            offset_lat = random.uniform(-0.02, 0.02)
            offset_lng = random.uniform(-0.02, 0.02)
            pharmacy_lat = lat + offset_lat
            pharmacy_lng = lng + offset_lng
            distance = calculate_distance(lat, lng, pharmacy_lat, pharmacy_lng)
            street = random.choice(street_names)
            address = f'{street}, {location_name}'
            name = pharmacy_names[i]
            
            pharmacies.append({
                'name': name,
                'address': address,
                'lat': pharmacy_lat,
                'lng': pharmacy_lng,
                # search_query is used by the frontend to build a correct Google Maps search URL
                'search_query': f'{name} near {location_name}',
                'distance': distance,
                'phone': f'+91-{random.randint(80, 99)}-{random.randint(10000000, 99999999)}',
                'open_now': random.choice(['Open 24 hours', 'Open', 'Open till 10 PM', 'Open till 11 PM']),
                'rating': f'{random.uniform(4.0, 4.9):.1f}',
                'type': 'pharmacy'
            })
        
        pharmacies.sort(key=lambda x: x['distance'])
        print(f"Found {len(pharmacies)} pharmacies near {location_name}")
        return pharmacies
        
    except Exception as e:
        print(f"Error finding pharmacies: {e}")
        return []
        
        # Check if response is valid
        if response.status_code != 200:
            print(f"Overpass API error: Status {response.status_code}")
            return []
        
        # Check if response has content
        if not response.text or response.text.strip() == '':
            print("Overpass API returned empty response")
            return []
        
        try:
            data = response.json()
        except ValueError as e:
            print(f"Error parsing JSON from Overpass API: {e}")
            print(f"Response text: {response.text[:200]}")
            return []
        
        pharmacies = []
        
        for element in data.get('elements', []):
            tags = element.get('tags', {})
            name = tags.get('name', 'Pharmacy')
            
            # Get coordinates
            if 'lat' in element and 'lon' in element:
                elem_lat = element['lat']
                elem_lng = element['lon']
            elif 'center' in element:
                elem_lat = element['center']['lat']
                elem_lng = element['center']['lon']
            else:
                continue
            
            # Calculate distance
            distance = calculate_distance(lat, lng, elem_lat, elem_lng)
            
            # Build address
            address_parts = []
            if tags.get('addr:housenumber'):
                address_parts.append(tags['addr:housenumber'])
            if tags.get('addr:street'):
                address_parts.append(tags['addr:street'])
            if tags.get('addr:city'):
                address_parts.append(tags['addr:city'])
            
            address = ', '.join(address_parts) if address_parts else 'Address not available'
            
            pharmacies.append({
                'name': name,
                'address': address,
                'distance': distance,
                'lat': elem_lat,
                'lng': elem_lng,
                'phone': tags.get('phone', tags.get('contact:phone', 'N/A')),
                'open_now': tags.get('opening_hours', 'Unknown'),
                'rating': 'N/A',
                'type': 'pharmacy'
            })
        
        # Sort by distance
        pharmacies.sort(key=lambda x: x['distance'])
        return pharmacies[:10]
        
    except Exception as e:
        print(f"Error finding pharmacies: {e}")
        return []


def find_nearby_diagnostic_centers(lat, lng, prescription='', radius=5000):
    """
    Find nearby diagnostic centers - Generates sample data near user's location
    """
    try:
        print(f"Searching for diagnostic centers near: {lat}, {lng} within {radius}m")
        
        # Use the shared reverse geocoding helper
        location_name = reverse_geocode_city(lat, lng)
        
        # Generate sample diagnostic centers near the user's location
        import random
        center_names = [
            'Manipal Hospital', 'Narayana Health', 'Fortis Hospital', 'Apollo Diagnostics',
            'Thyrocare Labs', 'Dr. Lal PathLabs', 'SRL Diagnostics', 'Vijaya Diagnostic',
            'Max Healthcare', 'Columbia Asia'
        ]
        
        areas = ['Main Road', 'Gandhi Nagar', 'City Center', 'Bus Stand Area', 'Railway Station Road',
                'Market Area', 'Hospital Road', 'College Road', 'Industrial Area', 'New Town']
        
        all_services = [
            ['CT Scan', 'MRI', 'X-Ray', 'Blood Tests', 'ECG'],
            ['Blood Tests', 'Ultrasound', 'X-Ray'],
            ['CT Scan', 'MRI', 'Blood Tests'],
            ['Blood Tests', 'Thyroid Test', 'Diabetes Test'],
            ['X-Ray', 'ECG', 'Blood Tests'],
            ['CT Scan', 'MRI', 'X-Ray', 'Ultrasound', 'Blood Tests']
        ]
        
        centers = []
        for i in range(6):
            offset_lat = random.uniform(-0.03, 0.03)
            offset_lng = random.uniform(-0.03, 0.03)
            center_lat = lat + offset_lat
            center_lng = lng + offset_lng
            distance = calculate_distance(lat, lng, center_lat, center_lng)
            area = random.choice(areas)
            address = f'{area}, {location_name}'
            name = center_names[i]
            
            centers.append({
                'name': name,
                'address': address,
                'lat': center_lat,
                'lng': center_lng,
                # search_query used by frontend for correct Google Maps URL
                'search_query': f'{name} near {location_name}',
                'distance': distance,
                'phone': f'+91-{random.randint(80, 99)}-{random.randint(10000000, 99999999)}',
                'open_now': random.choice(['Open 24 hours', 'Open', 'Open till 8 PM', 'Open till 9 PM']),
                'rating': f'{random.uniform(4.0, 4.9):.1f}',
                'services': random.choice(all_services),
                'type': 'diagnostic'
            })
        
        # Sort by distance
        centers.sort(key=lambda x: x['distance'])
        print(f"Found {len(centers)} diagnostic centers near {location_name}")
        return centers
        
    except Exception as e:
        print(f"Error finding diagnostic centers: {e}")
        return []
        
        # Check if response is valid
        if response.status_code != 200:
            print(f"Overpass API error: Status {response.status_code}")
            return []
        
        # Check if response has content
        if not response.text or response.text.strip() == '':
            print("Overpass API returned empty response")
            return []
        
        try:
            data = response.json()
        except ValueError as e:
            print(f"Error parsing JSON from Overpass API: {e}")
            print(f"Response text: {response.text[:200]}")
            return []
        
        centers = []
        
        for element in data.get('elements', []):
            tags = element.get('tags', {})
            name = tags.get('name', 'Diagnostic Center')
            
            # Get coordinates
            if 'lat' in element and 'lon' in element:
                elem_lat = element['lat']
                elem_lng = element['lon']
            elif 'center' in element:
                elem_lat = element['center']['lat']
                elem_lng = element['center']['lon']
            else:
                continue
            
            # Calculate distance
            distance = calculate_distance(lat, lng, elem_lat, elem_lng)
            
            # Build address
            address_parts = []
            if tags.get('addr:housenumber'):
                address_parts.append(tags['addr:housenumber'])
            if tags.get('addr:street'):
                address_parts.append(tags['addr:street'])
            if tags.get('addr:city'):
                address_parts.append(tags['addr:city'])
            
            address = ', '.join(address_parts) if address_parts else 'Address not available'
            
            # Determine services offered
            services = []
            name_lower = name.lower()
            if 'lab' in name_lower or 'pathology' in name_lower:
                services.append('Blood Tests')
            if 'scan' in name_lower or 'imaging' in name_lower or 'radiology' in name_lower:
                services.extend(['CT Scan', 'MRI', 'X-Ray'])
            if 'hospital' in name_lower:
                services.extend(['CT Scan', 'MRI', 'X-Ray', 'Blood Tests', 'ECG'])
            if not services:
                services = ['General Diagnostics']
            
            centers.append({
                'name': name,
                'address': address,
                'distance': distance,
                'lat': elem_lat,
                'lng': elem_lng,
                'phone': tags.get('phone', tags.get('contact:phone', 'N/A')),
                'open_now': tags.get('opening_hours', 'Unknown'),
                'rating': 'N/A',
                'services': services,
                'type': 'diagnostic'
            })
        
        # Sort by distance
        centers.sort(key=lambda x: x['distance'])
        return centers[:10]
        
    except Exception as e:
        print(f"Error finding diagnostic centers: {e}")
        return []


def calculate_distance(lat1, lng1, lat2, lng2):
    """
    Calculate distance between two coordinates using Haversine formula
    Returns distance in kilometers
    """
    import math
    
    R = 6371  # Earth's radius in kilometers
    
    lat1_rad = math.radians(lat1)
    lat2_rad = math.radians(lat2)
    delta_lat = math.radians(lat2 - lat1)
    delta_lng = math.radians(lng2 - lng1)
    
    a = math.sin(delta_lat/2)**2 + math.cos(lat1_rad) * math.cos(lat2_rad) * math.sin(delta_lng/2)**2
    c = 2 * math.atan2(math.sqrt(a), math.sqrt(1-a))
    
    distance = R * c
    return round(distance, 2)


# ==================== AUTO-TRIGGER NEARBY SERVICES ====================

@app.route('/doctor/complete-consultation/<int:consultation_id>', methods=['POST'])
@login_required
@role_required('doctor')
def complete_consultation_with_services(consultation_id):
    """
    Complete consultation and automatically trigger nearby services suggestion
    """
    diagnosis = request.form.get('diagnosis', '')
    prescription = request.form.get('prescription', '')
    
    conn = get_db()
    
    # Update consultation
    conn.execute('''
        UPDATE consultations 
        SET diagnosis = ?, prescription = ?, status = 'completed'
        WHERE id = ?
    ''', (diagnosis, prescription, consultation_id))
    
    # Get patient details
    consultation = conn.execute('''
        SELECT c.*, p.user_id as patient_user_id, u.name as patient_name
        FROM consultations c
        JOIN patients p ON c.patient_id = p.id
        JOIN users u ON p.user_id = u.id
        WHERE c.id = ?
    ''', (consultation_id,)).fetchone()
    
    if consultation:
        # Analyze prescription
        service_type = analyze_prescription_type(prescription)
        
        # Create notification for patient with nearby services link
        if service_type == 'pharmacy':
            message = f'Your consultation is complete. Dr. has prescribed medicines. Click here to find nearby medical stores.'
        elif service_type == 'diagnostic':
            message = f'Your consultation is complete. Dr. has recommended diagnostic tests. Click here to find nearby labs.'
        else:
            message = f'Your consultation is complete. Dr. has prescribed medicines and tests. Click here to find nearby services.'
        
        conn.execute('''
            INSERT INTO notifications (user_id, title, message, type)
            VALUES (?, ?, ?, ?)
        ''', (consultation['patient_user_id'], 
              '✅ Consultation Complete - Find Nearby Services', 
              message, 
              'consultation'))
    
    conn.commit()
    conn.close()
    
    flash('Consultation completed! Patient has been notified about nearby services.', 'success')
    return redirect(url_for('doctor_consultations'))


# ==================== DELETE CONSULTATION HISTORY ====================

@app.route('/patient/consultation/delete/<int:consultation_id>', methods=['POST'])
@login_required
@role_required('patient')
def delete_consultation_patient(consultation_id):
    """Delete a consultation from patient's history"""
    conn = get_db()
    
    # Verify the consultation belongs to this patient
    patient = conn.execute('SELECT id FROM patients WHERE user_id = ?', (session['user_id'],)).fetchone()
    consultation = conn.execute('SELECT * FROM consultations WHERE id = ? AND patient_id = ?', 
                                (consultation_id, patient['id'])).fetchone()
    
    if consultation:
        # Delete related feedback first
        conn.execute('DELETE FROM feedback WHERE consultation_id = ?', (consultation_id,))
        # Delete the consultation
        conn.execute('DELETE FROM consultations WHERE id = ?', (consultation_id,))
        conn.commit()
        conn.close()
        return jsonify({'success': True, 'message': 'Consultation deleted successfully'})
    else:
        conn.close()
        return jsonify({'success': False, 'message': 'Consultation not found or access denied'}), 403


@app.route('/doctor/consultation/delete/<int:consultation_id>', methods=['POST'])
@login_required
@role_required('doctor')
def delete_consultation_doctor(consultation_id):
    """Delete a consultation from doctor's records"""
    conn = get_db()
    
    # Verify the consultation belongs to this doctor
    doctor = conn.execute('SELECT id FROM doctors WHERE user_id = ?', (session['user_id'],)).fetchone()
    consultation = conn.execute('SELECT * FROM consultations WHERE id = ? AND doctor_id = ?', 
                                (consultation_id, doctor['id'])).fetchone()
    
    if consultation:
        # Delete related feedback first
        conn.execute('DELETE FROM feedback WHERE consultation_id = ?', (consultation_id,))
        # Delete the consultation
        conn.execute('DELETE FROM consultations WHERE id = ?', (consultation_id,))
        conn.commit()
        conn.close()
        return jsonify({'success': True, 'message': 'Consultation deleted successfully'})
    else:
        conn.close()
        return jsonify({'success': False, 'message': 'Consultation not found or access denied'}), 403


# ==================== DOCTOR RECOMMENDATIONS API ====================

@app.route('/api/recommend-doctors', methods=['POST'])
@login_required
@role_required('patient')
def recommend_doctors():
    """Recommend doctors based on predicted disease"""
    data = request.get_json()
    disease = data.get('disease', '').lower()
    
    # Enhanced disease to specialization mapping
    specialization_map = {
        'common cold': 'General Medicine',
        'flu': 'General Medicine',
        'fever': 'General Medicine',
        'cough': 'General Medicine',
        'headache': 'General Medicine',
        'diabetes': 'Endocrinologist',
        'hypertension': 'Cardiologist',
        'heart': 'Cardiologist',
        'cardiac': 'Cardiologist',
        'blood pressure': 'Cardiologist',
        'asthma': 'Pulmonologist',
        'pneumonia': 'Pulmonologist',
        'bronchitis': 'Pulmonologist',
        'lung': 'Pulmonologist',
        'respiratory': 'Pulmonologist',
        'arthritis': 'Rheumatologist',
        'joint': 'Rheumatologist',
        'bone': 'Orthopedist',
        'fracture': 'Orthopedist',
        'migraine': 'Neurologist',
        'seizure': 'Neurologist',
        'stroke': 'Neurologist',
        'brain': 'Neurologist',
        'depression': 'Psychiatrist',
        'anxiety': 'Psychiatrist',
        'mental': 'Psychiatrist',
        'stress': 'Psychiatrist',
        'skin': 'Dermatologist',
        'rash': 'Dermatologist',
        'acne': 'Dermatologist',
        'allergy': 'Allergist',
        'gastritis': 'Gastroenterologist',
        'ulcer': 'Gastroenterologist',
        'stomach': 'Gastroenterologist',
        'digestive': 'Gastroenterologist',
        'kidney': 'Nephrologist',
        'urine': 'Urologist',
        'eye': 'Ophthalmologist',
        'vision': 'Ophthalmologist',
        'ear': 'ENT Specialist',
        'throat': 'ENT Specialist',
        'nose': 'ENT Specialist',
    }
    
    # Find matching specialization with better matching
    recommended_spec = 'General Medicine'  # Default changed to match database
    max_matches = 0
    
    for keyword, spec in specialization_map.items():
        if keyword in disease:
            # Count how many words match for better accuracy
            matches = len([word for word in keyword.split() if word in disease])
            if matches > max_matches:
                max_matches = matches
                recommended_spec = spec
    
    # Get doctors with that specialization
    conn = get_db()
    doctors = conn.execute('''
        SELECT d.id, d.specialization, d.qualification, d.experience, d.consultation_fee,
               u.name, u.email, u.phone,
               COALESCE(AVG(f.rating), 4.0) as avg_rating,
               COUNT(DISTINCT f.id) as review_count,
               d.available
        FROM doctors d
        JOIN users u ON d.user_id = u.id
        LEFT JOIN feedback f ON d.id = f.doctor_id
        WHERE d.specialization LIKE ? AND d.available = 1 AND d.verification_status = 'Verified'
        GROUP BY d.id
        ORDER BY avg_rating DESC, d.experience DESC, review_count DESC
        LIMIT 8
    ''', (f'%{recommended_spec}%',)).fetchall()
    
    # If no specialists found, get general medicine doctors
    if not doctors:
        doctors = conn.execute('''
            SELECT d.id, d.specialization, d.qualification, d.experience, d.consultation_fee,
                   u.name, u.email, u.phone,
                   COALESCE(AVG(f.rating), 4.0) as avg_rating,
                   COUNT(DISTINCT f.id) as review_count,
                   d.available
            FROM doctors d
            JOIN users u ON d.user_id = u.id
            LEFT JOIN feedback f ON d.id = f.doctor_id
            WHERE d.specialization LIKE '%General%' AND d.available = 1 AND d.verification_status = 'Verified'
            GROUP BY d.id
            ORDER BY avg_rating DESC, d.experience DESC
            LIMIT 6
        ''').fetchall()
        recommended_spec = 'General Medicine'
    
    conn.close()
    
    doctor_list = []
    for doc in doctors:
        # Add additional doctor information for enhanced popups
        doctor_info = {
            'id': doc['id'],
            'name': doc['name'],
            'specialization': doc['specialization'],
            'qualification': doc['qualification'],
            'experience': doc['experience'],
            'fee': doc['consultation_fee'],
            'rating': round(doc['avg_rating'], 1),
            'reviews': doc['review_count'],
            'available': doc['available'],
            # Add computed fields for better UI
            'is_top_rated': doc['avg_rating'] >= 4.5,
            'is_experienced': doc['experience'] >= 10,
            'has_reviews': doc['review_count'] > 0,
            'fee_category': 'Premium' if doc['consultation_fee'] > 1000 else 'Standard' if doc['consultation_fee'] > 500 else 'Budget'
        }
        doctor_list.append(doctor_info)
    
    # Add recommendation confidence based on disease match
    confidence = 'High' if max_matches > 0 else 'Moderate'
    
    return jsonify({
        'recommended_specialization': recommended_spec,
        'doctors': doctor_list,
        'total_found': len(doctor_list),
        'confidence': confidence,
        'disease_analyzed': disease,
        'message': f'Found {len(doctor_list)} qualified {recommended_spec}s for your condition'
    })


# ==================== HELPER FUNCTIONS ====================

def analyze_prescription_type(prescription):
    """Analyze prescription to determine if it's pharmacy, diagnostic, or both"""
    prescription_lower = prescription.lower()
    
    # Keywords for pharmacy
    pharmacy_keywords = ['tablet', 'capsule', 'syrup', 'medicine', 'drug', 'medication', 
                         'antibiotic', 'painkiller', 'mg', 'ml', 'dose']
    
    # Keywords for diagnostic
    diagnostic_keywords = ['test', 'scan', 'x-ray', 'mri', 'ct', 'blood test', 'urine test',
                          'ecg', 'ultrasound', 'biopsy', 'lab', 'pathology']
    
    has_pharmacy = any(keyword in prescription_lower for keyword in pharmacy_keywords)
    has_diagnostic = any(keyword in prescription_lower for keyword in diagnostic_keywords)
    
    if has_pharmacy and has_diagnostic:
        return 'both'
    elif has_diagnostic:
        return 'diagnostic'
    else:
        return 'pharmacy'


def get_email_settings():
    """Get email settings from database"""
    try:
        conn = get_db()
        settings = conn.execute('SELECT * FROM email_settings WHERE id = 1').fetchone()
        conn.close()
        
        if settings and settings['mail_username']:
            return {
                'server': settings['mail_server'],
                'port': settings['mail_port'],
                'use_tls': bool(settings['mail_use_tls']),
                'username': settings['mail_username'],
                'password': settings['mail_password'],
                'sender_name': settings['mail_sender_name']
            }
        return None
    except:
        return None

# ==================== DOCTOR REGISTRATION API ====================

@app.route('/register-doctor', methods=['POST'])
def register_doctor_api():
    """
    API endpoint for doctor registration with document upload.
    Handles file uploads and creates pending verification account.
    """
    try:
        # Get form data (not JSON since we're using FormData for file upload)
        name = request.form.get('name', '').strip()
        email = request.form.get('email', '').strip().lower()
        password = request.form.get('password', '').strip()
        phone = request.form.get('phone', '').strip()
        specialization = request.form.get('specialization', '').strip()
        experience = request.form.get('experience', '0')
        consultation_fee = request.form.get('consultation_fee', '500')  # Get consultation fee
        
        # Get uploaded files
        degree_certificate = request.files.get('degree_certificate')
        council_registration = request.files.get('council_registration')
        
        # Validate required fields
        if not all([name, email, password, phone, specialization, consultation_fee]):
            return jsonify({
                'success': False,
                'message': 'All fields are required'
            }), 400
        
        # Validate files
        if not degree_certificate or not council_registration:
            return jsonify({
                'success': False,
                'message': 'Both documents (degree certificate and council registration) are required'
            }), 400
        
        # Basic validation
        if len(name) < 2:
            return jsonify({
                'success': False,
                'message': 'Name must be at least 2 characters long'
            }), 400
        
        if '@' not in email or '.' not in email:
            return jsonify({
                'success': False,
                'message': 'Invalid email format'
            }), 400
        
        if len(phone) < 10:
            return jsonify({
                'success': False,
                'message': 'Phone number must be at least 10 digits'
            }), 400
        
        if len(password) < 6:
            return jsonify({
                'success': False,
                'message': 'Password must be at least 6 characters long'
            }), 400
        
        # Validate file types
        allowed_extensions = {'pdf', 'jpg', 'jpeg', 'png'}
        
        def allowed_file(filename):
            return '.' in filename and filename.rsplit('.', 1)[1].lower() in allowed_extensions
        
        if not allowed_file(degree_certificate.filename):
            return jsonify({
                'success': False,
                'message': 'Degree certificate must be PDF, JPG, or PNG'
            }), 400
        
        if not allowed_file(council_registration.filename):
            return jsonify({
                'success': False,
                'message': 'Council registration must be PDF, JPG, or PNG'
            }), 400
        
        # Generate unique verification ID
        verification_id = generate_unique_doctor_verification_id()
        
        # Hash the provided password
        hashed_password = generate_password_hash(password)
        
        # Save uploaded files
        from werkzeug.utils import secure_filename
        upload_folder = 'uploads/doctor_documents'
        os.makedirs(upload_folder, exist_ok=True)
        
        # Create unique filenames
        degree_ext = degree_certificate.filename.rsplit('.', 1)[1].lower()
        council_ext = council_registration.filename.rsplit('.', 1)[1].lower()
        
        degree_filename = f"{verification_id}_degree.{degree_ext}"
        council_filename = f"{verification_id}_council.{council_ext}"
        
        degree_path = os.path.join(upload_folder, degree_filename)
        council_path = os.path.join(upload_folder, council_filename)
        
        degree_certificate.save(degree_path)
        council_registration.save(council_path)
        
        # Insert into database
        conn = get_db()
        
        try:
            c = conn.cursor()
            
            # Insert user with 'pending' status
            c.execute('''INSERT INTO users (name, email, password, role, phone, account_status)
                        VALUES (?, ?, ?, ?, ?, ?)''',
                     (name, email, hashed_password, 'doctor', phone, 'pending'))
            user_id = c.lastrowid
            
            # Insert doctor record with documents
            c.execute('''INSERT INTO doctors (user_id, specialization, qualification, experience, 
                                             consultation_fee, available, verification_id, verification_status,
                                             degree_certificate_path, council_registration_path, documents_uploaded)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''',
                     (user_id, specialization, 'Pending Verification', int(experience), 
                      int(consultation_fee), 0, verification_id, 'Pending',
                      degree_path, council_path, 1))
            
            # Create notification for doctor
            c.execute('''INSERT INTO notifications (user_id, title, message, type)
                        VALUES (?, ?, ?, ?)''',
                     (user_id, 
                      'Registration Submitted',
                      f'Your registration has been submitted for verification. Verification ID: {verification_id}. You will receive an email once verified.',
                      'registration'))
            
            # Notify admin about new registration
            admin_users = c.execute("SELECT id FROM users WHERE role = 'admin'").fetchall()
            for admin in admin_users:
                c.execute('''INSERT INTO notifications (user_id, title, message, type)
                            VALUES (?, ?, ?, ?)''',
                         (admin['id'],
                          '🔔 New Doctor Registration',
                          f'New doctor registration from {name} ({specialization}). Verification ID: {verification_id}. Please review documents.',
                          'admin'))
            
            conn.commit()
            
            return jsonify({
                'success': True,
                'message': 'Registration submitted successfully! Your account will be reviewed by admin.',
                'verification_id': verification_id,
                'status': 'Pending Verification',
                'note': 'You cannot login until your account is verified by admin. You will receive an email notification once verified.'
            }), 201
            
        except sqlite3.IntegrityError as e:
            # Delete uploaded files if database insert fails
            if os.path.exists(degree_path):
                os.remove(degree_path)
            if os.path.exists(council_path):
                os.remove(council_path)
            
            error_msg = str(e)
            if 'email' in error_msg.lower():
                return jsonify({
                    'success': False,
                    'message': 'Email already registered'
                }), 409
            elif 'phone' in error_msg.lower():
                return jsonify({
                    'success': False,
                    'message': 'Phone number already registered'
                }), 409
            else:
                return jsonify({
                    'success': False,
                    'message': 'Registration failed due to duplicate data'
                }), 409
        finally:
            conn.close()
            
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Server error: {str(e)}'
        }), 500

def generate_unique_doctor_verification_id():
    """Generate a unique verification ID in format DOC-2026-XXXX"""
    import random
    year = datetime.now().year
    
    conn = get_db()
    max_attempts = 100
    
    for _ in range(max_attempts):
        random_num = random.randint(1000, 9999)
        verification_id = f"DOC-{year}-{random_num}"
        
        # Check if this ID exists in doctors table (if column exists)
        try:
            existing = conn.execute('''SELECT COUNT(*) as count FROM doctors 
                                      WHERE verification_id = ?''', 
                                   (verification_id,)).fetchone()
            if existing['count'] == 0:
                conn.close()
                return verification_id
        except sqlite3.OperationalError:
            # Column doesn't exist yet, check notifications as fallback
            existing = conn.execute('''SELECT COUNT(*) as count FROM notifications 
                                      WHERE message LIKE ?''', 
                                   (f'%{verification_id}%',)).fetchone()
            if existing['count'] == 0:
                conn.close()
                return verification_id
    
    conn.close()
    # Fallback with timestamp if all random attempts fail
    import time
    return f"DOC-{year}-{int(time.time()) % 10000}"

# Doctor Registration Page (Frontend)
@app.route('/doctor-registration')
def doctor_registration_page():
    """Render the doctor registration form"""
    return render_template('doctor_registration.html')

# ==================== ADMIN DOCTOR VERIFICATION ====================

@app.route('/admin/verify-doctors')
@login_required
@role_required('admin')
def admin_verify_doctors():
    """Admin page to verify/reject doctor registrations"""
    filter_status = request.args.get('filter', 'pending')
    
    conn = get_db()
    
    # Get statistics
    stats = {
        'pending': conn.execute('''SELECT COUNT(*) as count FROM doctors 
                                  WHERE verification_status = "Pending"''').fetchone()['count'],
        'verified': conn.execute('''SELECT COUNT(*) as count FROM doctors 
                                   WHERE verification_status = "Verified"''').fetchone()['count'],
        'rejected': conn.execute('''SELECT COUNT(*) as count FROM doctors 
                                   WHERE verification_status = "Rejected"''').fetchone()['count'],
        'total': conn.execute('SELECT COUNT(*) as count FROM doctors').fetchone()['count']
    }
    
    # Build query based on filter
    query = '''SELECT d.id as doctor_id, d.*, u.name, u.email, u.phone, u.created_at
               FROM doctors d
               JOIN users u ON d.user_id = u.id
               WHERE 1=1'''
    
    if filter_status == 'pending':
        query += ' AND d.verification_status = "Pending"'
    elif filter_status == 'verified':
        query += ' AND d.verification_status = "Verified"'
    elif filter_status == 'rejected':
        query += ' AND d.verification_status = "Rejected"'
    # 'all' shows everything
    
    query += ' ORDER BY u.created_at DESC'
    
    doctors = conn.execute(query).fetchall()
    conn.close()
    
    return render_template('admin_verify_doctors.html', 
                         doctors=doctors, 
                         stats=stats, 
                         filter=filter_status)

@app.route('/admin/verify-doctor/<int:doctor_id>', methods=['POST'])
@login_required
@role_required('admin')
def admin_verify_doctor(doctor_id):
    """Verify or reject a doctor"""
    action = request.form.get('action')  # 'verify' or 'reject'
    notes = request.form.get('notes', '').strip()
    
    conn = get_db()
    
    if action == 'verify':
        # Verify the doctor
        conn.execute('''UPDATE doctors 
                       SET verification_status = "Verified", 
                           available = 1,
                           qualification = CASE 
                               WHEN qualification = "Pending Verification" THEN "MBBS" 
                               ELSE qualification 
                           END,
                           verification_notes = ?
                       WHERE id = ?''', (notes, doctor_id))
        
        # Get doctor details for notification
        doctor = conn.execute('''SELECT d.user_id, u.name, u.email
                                FROM doctors d 
                                JOIN users u ON d.user_id = u.id 
                                WHERE d.id = ?''', (doctor_id,)).fetchone()
        
        # Update user account status to 'active'
        conn.execute('''UPDATE users SET account_status = 'active' WHERE id = ?''', (doctor['user_id'],))
        
        # Send notification to doctor
        conn.execute('''INSERT INTO notifications (user_id, title, message, type)
                       VALUES (?, ?, ?, ?)''',
                    (doctor['user_id'],
                     '✅ Account Verified',
                     'Congratulations! Your doctor account has been verified by admin. You can now login and accept appointments.',
                     'verification'))
        
        # Send email notification to doctor
        try:
            # Get doctor's verification ID
            doctor_details = conn.execute('''SELECT verification_id FROM doctors WHERE id = ?''', (doctor_id,)).fetchone()
            verification_id = doctor_details['verification_id'] if doctor_details else f"DOC-{doctor_id}"
            
            send_doctor_verification_email(
                doctor_email=doctor['email'],
                doctor_name=doctor['name'],
                verification_id=verification_id,
                status='approved',
                notes=notes
            )
            print(f"[Email] ✅ Verification approval email sent to Dr. {doctor['name']}")
        except Exception as e:
            print(f"[Email] ❌ Failed to send verification email to Dr. {doctor['name']}: {str(e)}")
        
        conn.commit()
        flash(f'✅ Dr. {doctor["name"]} has been verified successfully! They can now login.', 'success')
        
    elif action == 'reject':
        rejection_reason = request.form.get('rejection_reason', 'Documents did not meet verification requirements')
        
        # Reject the doctor
        conn.execute('''UPDATE doctors 
                       SET verification_status = "Rejected", 
                           available = 0,
                           verification_notes = ?
                       WHERE id = ?''', (f"Rejected: {rejection_reason}", doctor_id))
        
        # Get doctor details for notification
        doctor = conn.execute('''SELECT d.user_id, u.name, u.email
                                FROM doctors d 
                                JOIN users u ON d.user_id = u.id 
                                WHERE d.id = ?''', (doctor_id,)).fetchone()
        
        # Update user account status to 'rejected'
        conn.execute('''UPDATE users SET account_status = 'rejected' WHERE id = ?''', (doctor['user_id'],))
        
        # Send notification to doctor
        conn.execute('''INSERT INTO notifications (user_id, title, message, type)
                       VALUES (?, ?, ?, ?)''',
                    (doctor['user_id'],
                     '❌ Account Verification Rejected',
                     f'Your doctor account verification was not approved. Reason: {rejection_reason}. Please contact admin for more information.',
                     'verification'))
        
        # Send email notification to doctor
        try:
            # Get doctor's verification ID
            doctor_details = conn.execute('''SELECT verification_id FROM doctors WHERE id = ?''', (doctor_id,)).fetchone()
            verification_id = doctor_details['verification_id'] if doctor_details else f"DOC-{doctor_id}"
            
            send_doctor_verification_email(
                doctor_email=doctor['email'],
                doctor_name=doctor['name'],
                verification_id=verification_id,
                status='rejected',
                notes=rejection_reason
            )
            print(f"[Email] ✅ Verification rejection email sent to Dr. {doctor['name']}")
        except Exception as e:
            print(f"[Email] ❌ Failed to send verification email to Dr. {doctor['name']}: {str(e)}")
        
        conn.commit()
        flash(f'❌ Dr. {doctor["name"]} has been rejected.', 'warning')
    
    conn.close()
    return redirect(url_for('admin_verify_doctors'))

@app.route('/admin/view-document/<int:doctor_id>/<doc_type>')
@login_required
@role_required('admin')
def admin_view_document(doctor_id, doc_type):
    """View or download doctor documents"""
    from flask import send_file
    
    conn = get_db()
    doctor = conn.execute('SELECT * FROM doctors WHERE id = ?', (doctor_id,)).fetchone()
    conn.close()
    
    if not doctor:
        flash('Doctor not found', 'danger')
        return redirect(url_for('admin_verify_doctors'))
    
    # Get document path based on type
    if doc_type == 'degree':
        file_path = doctor['degree_certificate_path'] if 'degree_certificate_path' in doctor.keys() else None
    elif doc_type == 'council':
        file_path = doctor['council_registration_path'] if 'council_registration_path' in doctor.keys() else None
    else:
        flash('Invalid document type', 'danger')
        return redirect(url_for('admin_verify_doctors'))
    
    if not file_path or not os.path.exists(file_path):
        flash('Document not found', 'danger')
        return redirect(url_for('admin_verify_doctors'))
    
    # Send file for viewing/download
    return send_file(file_path, as_attachment=False)


if __name__ == '__main__':
    init_db()
    
    # Start appointment reminder scheduler
    if REMINDER_SCHEDULER_AVAILABLE:
        start_reminder_scheduler()
    
    if SOCKETIO_AVAILABLE and socketio:
        socketio.run(app, debug=True, port=5000, allow_unsafe_werkzeug=True)
    else:
        app.run(debug=True, port=5000)
